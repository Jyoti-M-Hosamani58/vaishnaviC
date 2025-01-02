import datetime
import json
import os
import random
from datetime import datetime
from datetime import timedelta

from consign.settings import BASE_DIR
from consign_app.models import (Login, Collection, DriverLocation, UserLogin, AddConsignment, AddConsignmentTemp, Disel,
                                AddTrack, FeedBack, Branch, Driver, Vehicle, Staff, Consignee, Consignor,
                                TripSheetTemp, TripSheetPrem, Account, Expenses, Transporter, Fullload,Balance,Fullloadexpenses)
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db.models import Count
from django.db.models.fields import return_None
from django.http import JsonResponse
from django.shortcuts import render, reverse, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST


# from django.core.mail import send_mail

# import datetime
# from .models import AddTrack, AddConsignment


# Initialize Razorpay client with your API keys
from django.views.generic import TemplateView
class OfflineView(TemplateView):
    template_name = 'offline.html'  #

# Create your views here.
def index(request):
    return render(request, 'index.html')


def feedback(request):
    uid = request.session.get('username')
    if not uid:
        return redirect('login')  # Redirect to login if session does not have username

    # Fetch only the receiver_email column
    userdata = AddConsignment.objects.filter(receiver_email=uid).values_list('receiver_email', flat=True)

    if request.method == "POST":
        feed = request.POST.get('feedback')

        if userdata.exists():
            username = userdata[0]  # Extract the first email from the list

            FeedBack.objects.create(
                username=username,
                feedback=feed
            )
            messages.success(request, 'Feedback sent successfully')
            return redirect('feedback')
        else:
            messages.error(request, 'User not found')
            return render(request, 'feedback.html')

    return render(request, 'feedback.html')


def view_feedback(request):
    userdata = FeedBack.objects.all()
    return render(request, 'view_feedback.html', {'userdata': userdata})


def staff_home(request):
    return render(request, 'staff_home.html')


def staff_nav(request):
    return render(request, 'staff_nav.html')


def index_menu(request):
    return render(request, 'index_menu.html')


def admin_home(request):
    return render(request, 'admin_home.html')


def user_home(request):
    return render(request, 'user_home.html')


def user_home(request):
    return render(request, 'user_home.html')


def user_menu(request):
    return render(request, 'user_menu.html')


def nav(request):
    return render(request, 'nav.html')


def branch_home(request):
    return render(request, 'branch_home.html')


def staff_home(request):
    return render(request, 'staff_home.html')


def userlogin(request):
    if request.method == "POST":
        username = request.POST.get('t1')
        password = request.POST.get('t2')
        request.session['username'] = username
        ucount = Login.objects.filter(username=username).count()
        if ucount >= 1:
            udata = Login.objects.get(username=username)
            upass = udata.password
            utype = udata.utype
            if password == upass:
                request.session['utype'] = utype
                if utype == 'user':
                    return render(request, 'user_home.html')
                if utype == 'admin':
                    return render(request, 'admin_home.html')
                if utype == 'branch':
                    return render(request, 'branch_home.html')
                if utype == 'staff':
                    return render(request, 'staff_home.html')
                if utype == 'driver':
                    return redirect('location')
            else:
                return render(request, 'userlogin.html', {'msg': 'Invalid Password'})
        else:
            return render(request, 'userlogin.html', {'msg': 'Invalid Username'})
    return render(request, 'userlogin.html')


from django.db.models import Max


def addConsignment(request):
    if request.method == "POST":
        now = datetime.now().replace(microsecond=0)
        con_date = now.strftime("%Y-%m-%d")
        current_time = now.strftime("%H:%M:%S")

        uid = request.session.get('username')
        branch = Staff.objects.get(staffPhone=uid)
        uname = branch.Branch
        branchemail = branch.branchemail
        username = branch.staffname

        # Get the last track_id and increment it
        last_track_id = AddConsignment.objects.aggregate(Max('track_id'))['track_id__max']
        track_id = int(last_track_id) + 1 if last_track_id else 1000
        con_id = str(track_id)

        # Get the last Consignment_id and increment it
        last_con_id = AddConsignment.objects.aggregate(Max('Consignment_id'))['Consignment_id__max']
        Consignment_id = last_con_id + 1 if last_con_id else 1000
        Consignment_id = str(Consignment_id)

        last_cust_id = Consignee.objects.aggregate(Max('cust_id'))['cust_id__max']
        cust_id = last_cust_id + 1 if last_cust_id else 1
        cust_id = str(cust_id)

        last_consignor_id = Consignor.objects.aggregate(Max('cust_id'))['cust_id__max']
        consignor_id = last_consignor_id + 1 if last_consignor_id else 1000
        consignor_id = str(consignor_id)

        # Sender details
        send_name = request.POST.get('a1')
        send_mobile = request.POST.get('a2')
        send_address = request.POST.get('a4')
        sender_GST = request.POST.get('sendergst')

        # Receiver details
        rec_name = request.POST.get('a5')
        rec_mobile = request.POST.get('a6')
        rec_address = request.POST.get('a8')
        rec_GST = request.POST.get('receivergst')

        # Check if the Consignor already exists
        consignor, created = Consignor.objects.get_or_create(
            sender_name=send_name,
            defaults={
                'sender_mobile': send_mobile,
                'sender_address': send_address,
                'sender_GST': sender_GST,
                'branch': uname,
                'cust_id': consignor_id  # Temporarily set cust_id to None
            }
        )

        # If the consignor was created, assign a new cust_id
        if created:
            consignor_id = consignor.cust_id  # Use the newly created cust_id
        else:
            consignor_id = consignor.cust_id  # Retain existing cust_id if not created

        # Save consignor mobile number in UserLogin table
        UserLogin.objects.update_or_create(
            email=send_mobile,  # Use sender_mobile as the unique key
            defaults={
                'username': send_name,  # Set username to sender_name
                # Add any other fields as necessary
            }
        )

        # Check if the Consignee already exists
        consignee, created = Consignee.objects.get_or_create(
            receiver_name=rec_name,
            defaults={
                'receiver_mobile': rec_mobile,
                'receiver_address': rec_address,
                'receiver_GST': rec_GST,
                'branch': uname,
                'cust_id': cust_id  # Temporarily set cust_id to None
            }
        )

        # Save consignee mobile number in UserLogin table
        UserLogin.objects.update_or_create(
            email=rec_mobile,  # Use receiver_mobile as the unique key
            defaults={
                'username': rec_name,  # Set username to receiver_name
                # Add any other fields as necessary
            }
        )

        # If the consignee was created, assign a new cust_id
        if created:
            cust_id = consignee.cust_id  # Use the newly created cust_id
        else:
            cust_id = consignee.cust_id  # Retain existing cust_id if not created

        # Handling copies
        copies = []
        if request.POST.get('consignor_copy'):
            copies.append('Consignor Copy')
        if request.POST.get('consignee_copy'):
            copies.append('Consignee Copy')
        if request.POST.get('lorry_copy'):
            copies.append('Lorry Copy')
        copy_type = ', '.join(copies)  # Combine into a single string

        # Handling product entries
        products = request.POST.getlist('product[]')
        pieces = request.POST.getlist('pieces[]')

        # Other consignment details
        delivery = request.POST.get('delivery_option')
        prod_invoice = request.POST.get('prod_invoice')
        prod_price = request.POST.get('prod_price')
        weight = float(request.POST.get('weight') or 0)
        weightAmt = float(request.POST.get('weightAmt') or 0)
        freight = float(request.POST.get('freight') or 0)
        hamali = float(request.POST.get('hamali') or 0)
        door_charge = float(request.POST.get('door_charge') or 0)
        st_charge = float(request.POST.get('st_charge') or 0)
        cost = float(request.POST.get('cost') or 0)
        bal = float(request.POST.get('balance') or 0)
        pay_status = request.POST.get('payment')
        route_from = request.POST.get('from')
        route_to = request.POST.get('to')
        eway_bill = request.POST.get('ewaybill_no')

        utype = request.session.get('utype')
        branch_value = 'admin' if utype == 'admin' else uname

        # Determine the appropriate name based on pay_status
        if pay_status == 'Consigner_AC':
            account_name = send_name

        elif pay_status == 'Consignee_AC':
            account_name = rec_name
        else:
            account_name = send_name  # Default to sender_name if pay_status is neither

        # Create consignment records
        for product, piece in zip(products, pieces):
            if not product or not piece:
                continue
            AddConsignment.objects.create(
                track_id=con_id,
                Consignment_id=Consignment_id,
                sender_name=send_name,
                sender_mobile=send_mobile,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_mobile=rec_mobile,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                desc_product=product,
                pieces=piece,
                prod_invoice=prod_invoice,
                prod_price=prod_price,
                weightAmt=weightAmt,
                weight=weight,
                balance=bal,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total_cost=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                time=current_time,
                copy_type=copy_type,
                delivery=delivery,
                eway_bill=eway_bill,
                consignment_status='Pending',
                branchemail=branchemail,
                status='Consignment Added'

            )
            AddConsignmentTemp.objects.create(
                track_id=con_id,
                Consignment_id=Consignment_id,
                sender_name=send_name,
                sender_mobile=send_mobile,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_mobile=rec_mobile,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                desc_product=product,
                pieces=piece,
                prod_invoice=prod_invoice,
                prod_price=prod_price,
                weightAmt=weightAmt,
                weight=weight,
                balance=bal,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total_cost=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                time=current_time,
                copy_type=copy_type,
                delivery=delivery,
                eway_bill=eway_bill
            )
            Collection.objects.create(
                lrNo=con_id,
                sender_name=send_name,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                weightRate=weightAmt,
                weight=weight,
                balance=cost,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                consignment_status='Pending',
                amount=0,

            )

        if pay_status in ['Consigner_AC', 'Consignee_AC']:
            try:
                # Confirm this block executes
                print(f"Attempting to save to Account with track_id {con_id}, sender_name {account_name}, cost {cost}")

                previous_balance_entry = Account.objects.filter(sender_name=account_name).order_by('-Date').first()
                previous_balance = float(previous_balance_entry.Balance) if previous_balance_entry else 0.0
                updated_balance = previous_balance + cost

                # Proceed to save/update account data
                Account.objects.update_or_create(
                    track_number=con_id,
                    defaults={
                        'sender_name': account_name,
                        'Date': now,
                        'particulars': f'LR: {con_id}',
                        'debit': cost,
                        'credit': 0,
                        'TrType': 'sal',
                        'Balance': updated_balance,
                        'headname': 'Account Receivable',
                        'Branch': uname
                    }
                )
                print(f"Successfully updated balance for {account_name}: {updated_balance}")
            except Exception as e:
                print(f"Error while saving to Account table: {e}")

        return redirect('printConsignment', track_id=con_id)
    return render(request, 'addConsignment.html')


def printConsignment(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity

    try:
        consignments = AddConsignment.objects.filter(track_id=track_id)
        uid = request.session.get('username')

        staff = Staff.objects.get(staffPhone=uid)
        user_branch = staff.Branch  # Adjust if the branch info is stored differently
        branchemail = staff.branchemail
        branchdetails = Branch.objects.get(email=branchemail)

        if not consignments.exists():
            return render(request, '404.html')  # Handle case where no consignments are found.
        branchdetails.services = mark_safe(branchdetails.services.replace(',', '<br/>'))

        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'desc_product': consignment.desc_product,

            }
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)
            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    return render(request, 'printConsignment.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),  # Include the aggregated copy types
        'totalqty': totalqty  # Pass total quantity to the template

    })


def invoiceConsignment(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity

    try:
        # Filter consignments by track_id
        consignments = AddConsignment.objects.filter(track_id=track_id)
        # Get common details from the first consignment
        consignment = consignments.first()

        # Fetch the branch name from the consignment
        branch_name = consignment.branch  # Adjust this field based on your model
        branchemail = consignment.branchemail
        # Fetch branch details using the branch name
        branchdetails = get_object_or_404(Branch, companyname=branchemail)

        if not consignments.exists():
            return render(request, '404.html')  # Handle case where no consignments are found.
        branchdetails.services = mark_safe(branchdetails.services.replace(',', '<br/>'))

        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'desc_product': consignment.desc_product,

            }
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)
            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    return render(request, 'invoiceConsignment.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),  # Include the aggregated copy types
        'totalqty': totalqty  # Pass total quantity to the template

    })


def view_consignment(request):
    uid = request.session.get('username')
    grouped_userdata = {}
    highlight_ids = []  # To store track IDs to be highlighted

    if uid:
        try:
            from_date_str = request.POST.get('from_date')
            to_date_str = request.POST.get('to_date')

            consigner = request.POST.get('consigner')
            consignee = request.POST.get('consignee')
            track_id = request.POST.get('lrno')

            # Parse dates
            from_date = parse_date(from_date_str) if from_date_str else None
            to_date = parse_date(to_date_str) if to_date_str else None

            # Fetch the staff and associated branch
            staff = Staff.objects.get(staffPhone=uid)
            user_branch = staff.Branch

            # Start building the query
            consignments = AddConsignment.objects.filter(branch=user_branch)

            if consigner:
                consignments = consignments.filter(sender_name=consigner)
            if consignee:
                consignments = consignments.filter(receiver_name=consignee)
            if track_id:
                consignments = consignments.filter(track_id=track_id)

            if from_date and to_date:
                consignments = consignments.filter(date__range=(from_date, to_date))
            elif from_date:
                consignments = consignments.filter(date__gte=from_date)
            elif to_date:
                consignments = consignments.filter(date__lte=to_date)

            # Group consignments by track_id and concatenate product details
            for consignment in consignments:
                track_id = str(consignment.track_id)  # Ensure track_id is always a string
                if track_id not in grouped_userdata:
                    grouped_userdata[track_id] = {
                        'route_from': consignment.route_from,
                        'route_to': consignment.route_to,
                        'sender_name': consignment.sender_name,
                        'sender_mobile': consignment.sender_mobile,
                        'receiver_name': consignment.receiver_name,
                        'receiver_mobile': consignment.receiver_mobile,
                        'total_cost': 0,
                        'pieces': 0,
                        'weight': consignment.weight,
                        'pay_status': consignment.pay_status,
                        'products': []
                    }
                # Aggregate total cost and pieces
                grouped_userdata[track_id]['total_cost'] += consignment.total_cost
                grouped_userdata[track_id]['pieces'] += consignment.pieces

                # Concatenate product details without ID
                product_detail = consignment.desc_product
                grouped_userdata[track_id]['products'].append(product_detail)

            # Identify which track_ids are present in TripSheetPrem as LRno
            lrnos_in_tripsheet = TripSheetPrem.objects.filter(
                LRno__in=grouped_userdata.keys()
            ).values_list('LRno', flat=True)

            # Normalize highlight_ids to strings
            highlight_ids = [str(lrno).strip() for lrno in lrnos_in_tripsheet]

        except ObjectDoesNotExist:
            grouped_userdata = {}


    # Convert the list of product details to a single string
    for track_id, details in grouped_userdata.items():
        details['products'] = ', '.join(details['products'])

    print("Highlight IDs:", highlight_ids)
    print("Grouped User Data Keys:", list(grouped_userdata.keys()))

    return render(request, 'view_consignment.html', {
        'grouped_userdata': grouped_userdata,
        'highlight_ids': highlight_ids
    })

def user_view_consignment(request):
    uid = request.session['username']
    userdata = AddConsignment.objects.filter(receiver_email=uid).values()
    return render(request, 'user_view_consignment.html', {'userdata': userdata})


def consignment_edit(request, pk):
    userdata = AddConsignment.objects.filter(id=pk).first()  # Retrieve a single object or None

    if request.method == "POST":
        track_id = userdata.track_id
        con_date = userdata.date

        send_name = request.POST.get('a1')
        send_mobile = request.POST.get('a2')
        send_email = request.POST.get('a3')
        send_address = request.POST.get('a4')

        rec_name = request.POST.get('a5')
        rec_mobile = request.POST.get('a6')
        rec_email = request.POST.get('a7')
        rec_address = request.POST.get('a8')

        cost = request.POST.get('a9')

        # Update the object
        userdata.track_no = track_id
        userdata.sender_name = send_name
        userdata.sender_mobile = send_mobile
        userdata.sender_email = send_email
        userdata.sender_address = send_address
        userdata.receiver_name = rec_name
        userdata.receiver_mobile = rec_mobile
        userdata.receiver_email = rec_email
        userdata.receiver_address = rec_address
        userdata.total_cost = cost
        userdata.date = con_date
        userdata.save()

        # Redirect to a different URL after successful update
        base_url = reverse('view_consignment')
        return redirect(base_url)

    return render(request, 'consignment_edit.html', {'userdata': userdata})


def consignment_delete(request, pk):
    udata = AddConsignment.objects.get(id=pk)
    udata.delete()
    base_url = reverse('view_consignment')
    return redirect(base_url)


def addTrack(request):
    consignments = AddConsignment.objects.all().order_by('-id')  # Fetch all consignments ordered by id descending
    if request.method == "POST":
        now = datetime.datetime.now()
        con_date = now.strftime("%Y-%m-%d")

        track_id = request.POST.get('a1')
        status = request.POST.get('status')  # Retrieve status from the form

        # Retrieve total_cost from AddConsignment table based on some condition
        # For example, you can get it based on track_id or any other criteria

        # If the selected status is "Other", retrieve the custom status from the form
        if status == "Other":
            custom_status = request.POST.get('a2')
        else:
            custom_status = None

        # Create AddTrack object with retrieved total_cost
        AddTrack.objects.create(
            track_id=track_id,
            description=status,
            date=con_date

        )

        return render(request, 'addTrack.html', {'msg': 'Added'})
    return render(request, 'addTrack.html', {'consignments': consignments})


def search_results(request):
    tracker_id = request.GET.get('tracker_id')
    consignments = AddConsignment.objects.all().order_by('-id')  # Fetch all consignment data

    if tracker_id:
        try:
            trackers = AddTrack.objects.filter(track_id=tracker_id)
            if trackers.exists():
                return render(request, 'search_results.html', {'trackers': trackers, 'consignments': consignments})
            else:
                message = f"No tracking information found for ID: {tracker_id}"
                return render(request, 'search_results.html', {'message': message, 'consignments': consignments})
        except Exception as e:
            message = f"Error occurred: {str(e)}"
            return render(request, 'search_results.html', {'message': message, 'consignments': consignments})
    else:
        return render(request, 'search_results.html',
                      {'message': "Please enter a tracker ID.", 'consignments': consignments})


def track_delete(request, pk):
    udata = AddTrack.objects.get(id=pk)
    udata.delete()
    base_url = reverse('search_results')
    return redirect(base_url)


def user_search_results(request):
    tracker_id = request.GET.get('tracker_id')

    if tracker_id:
        try:
            trackers = AddTrack.objects.filter(track_id=tracker_id)
            if trackers.exists():
                return render(request, 'user_search_results.html', {'trackers': trackers})
            else:
                message = f"No tracking information found for ID: {tracker_id}"
                return render(request, 'user_search_results.html', {'message': message})
        except Exception as e:
            message = f"Error occurred: {str(e)}"
            return render(request, 'user_search_results.html', {'message': message})
    else:
        return render(request, 'user_search_results.html', {'message': "Please enter a tracker ID."})


def branch(request):
    if request.method == "POST":
        companyname = request.POST.get('companyname')
        headname = request.POST.get('headname')
        phonenumber = request.POST.get('phonenumber')
        email = request.POST.get('email')
        password = request.POST.get('password')
        gst = request.POST.get('gst')
        address = request.POST.get('address')
        services = request.POST.get('services')
        agency = request.POST.get('agency')

        utype = 'branch'

        if Login.objects.filter(username=email).exists():
            messages.error(request, 'Username (email) already exists.')
            return render(request, 'branch.html')

        # If the email does not exist, create the branch and login records
        Branch.objects.create(
            companyname=companyname,
            phonenumber=phonenumber,
            email=email,
            gst=gst,
            address=address,
            services=services,
            agency=agency,
            headname=headname,
            password=password
        )
        Login.objects.create(utype=utype, username=email, password=password, name=headname)

        messages.success(request, 'Branch created successfully.')

    return render(request, 'branch.html')


def view_branch(request):
    data = Branch.objects.all()
    return render(request, 'view_branch.html', {'data': data})


def edit_branch(request, pk):
    data = Branch.objects.filter(id=pk).first()  # Retrieve a single object or None

    original_email = data.email

    if request.method == "POST":
        companyname = request.POST.get('companyname')
        headname = request.POST.get('headname')
        phonenumber = request.POST.get('phonenumber')
        email = request.POST.get('email')
        gst = request.POST.get('gst')
        address = request.POST.get('address')
        services = request.POST.get('services')
        agency = request.POST.get('agency')
        password = request.POST.get('password')

        # Update the object
        data.companyname = companyname
        data.headname = headname
        data.phonenumber = phonenumber
        data.email = email
        data.gst = gst
        data.address = address
        data.services = services
        data.agency = agency
        data.password = password
        data.save()

        # Update the Login record using the original staffPhone
        user = Login.objects.filter(username=original_email).first()  # Fetch the user with the original phone number
        if user:
            user.username = email  # Update username to the new phone number
            user.name = headname  # Update name
            user.password = password
            user.save()
        # Redirect to a different URL after successful update
        base_url = reverse('view_branch')
        return redirect(base_url)

    return render(request, 'edit_branch.html', {'data': data})


def branch_delete(request, pk):
    udata = Branch.objects.get(id=pk)
    user = Login.objects.filter(username=udata.email).first()
    if user:
        user.delete()
    udata.delete()
    base_url = reverse('view_branch')
    return redirect(base_url)


def driver(request):
    if request.method == "POST":
        driver_name = request.POST.get('driver_name')
        phone_number = request.POST.get('phonenumber')
        address = request.POST.get('address')
        passport = request.POST.get('passport')
        license = request.POST.get('license')
        aadhar = request.POST.get('aadhar')

        passportfile = request.FILES['passport']
        fs = FileSystemStorage()
        filepassport = fs.save(passportfile.name, passportfile)
        upload_file_url = fs.url(filepassport)
        path = os.path.join(BASE_DIR, '/media/' + filepassport)

        licensefile = request.FILES['license']
        fs = FileSystemStorage()
        filelicense = fs.save(licensefile.name, licensefile)
        upload_file_url = fs.url(filelicense)
        path = os.path.join(BASE_DIR, '/media/' + filelicense)

        aadharfile = request.FILES['aadhar']
        fs = FileSystemStorage()
        fileaadhar = fs.save(aadharfile.name, aadharfile)
        upload_file_url = fs.url(fileaadhar)
        path = os.path.join(BASE_DIR, '/media/' + fileaadhar)

        Driver.objects.create(
            driver_name=driver_name,
            phone_number=phone_number,
            address=address,
            passport=passportfile,
            license=licensefile,
            aadhar=aadharfile,
            location_sharing_active='True'
        )
        Login.objects.create(
            username=phone_number,
            password=phone_number,
            utype='driver'
        )
    return render(request, 'driver.html')


def view_driver(request):
    data = Driver.objects.all()
    return render(request, 'view_driver.html', {'data': data})


def driver_edit(request, pk):
    data = Driver.objects.filter(id=pk).first()  # Retrieve a single object or None

    if request.method == "POST":
        driver_name = request.POST.get('driver_name')
        phone_number = request.POST.get('phone_number')
        address = request.POST.get('address')

        # Update the object
        data.driver_name = driver_name
        data.phone_number = phone_number
        data.address = address

        data.save()

        # Redirect to a different URL after successful update
        base_url = reverse('view_driver')
        return redirect(base_url)

    return render(request, 'driver_edit.html', {'data': data})


def driver_delete(request, pk):
    udata = Driver.objects.get(id=pk)
    udata.delete()
    base_url = reverse('view_driver')
    return redirect(base_url)


def vehicle(request):
    if request.method == "POST":
        vehicle_number = request.POST.get('vehicle_number')
        rcdate = request.POST.get('rcdate')
        incurencedate = request.POST.get('incurencedate')
        permitdate = request.POST.get('permitdate')
        taxdate = request.POST.get('taxdate')
        emissiondate = request.POST.get('emissiondate')

        rcfile = request.FILES['rc']
        fs = FileSystemStorage()
        filerc = fs.save(rcfile.name, rcfile)
        upload_file_url = fs.url(filerc)
        path = os.path.join(BASE_DIR, '/media/' + filerc)

        incurencefile = request.FILES['incurence']
        fs = FileSystemStorage()
        fileincurence = fs.save(incurencefile.name, incurencefile)
        upload_file_url = fs.url(fileincurence)
        path = os.path.join(BASE_DIR, '/media/' + fileincurence)

        permitfile = request.FILES['permit']
        fs = FileSystemStorage()
        filepermit = fs.save(permitfile.name, permitfile)
        upload_file_url = fs.url(filepermit)
        path = os.path.join(BASE_DIR, '/media/' + filepermit)

        taxfile = request.FILES['tax']
        fs = FileSystemStorage()
        filetax = fs.save(taxfile.name, taxfile)
        upload_file_url = fs.url(filetax)
        path = os.path.join(BASE_DIR, '/media/' + filetax)

        emissionfile = request.FILES['emission']
        fs = FileSystemStorage()
        fileemission = fs.save(emissionfile.name, emissionfile)
        upload_file_url = fs.url(fileemission)
        path = os.path.join(BASE_DIR, '/media/' + fileemission)

        if Vehicle.objects.filter(vehicle_number=vehicle_number).exists():
            messages.error(request, 'vehicle number already exists.')
            return render(request, 'vehicle.html')

        Vehicle.objects.create(
            vehicle_number=vehicle_number,
            rccard=rcfile,
            rccardate=rcdate,
            incurencedate=incurencedate,
            incurence=incurencefile,
            permit=permitfile,
            permitdate=permitdate,
            tax=taxfile,
            taxdate=taxdate,
            emission=emissionfile,
            emissiondate=emissiondate
        )
        messages.success(request, 'Vehicle created successfully.')
    return render(request, 'vehicle.html')


from django.utils.timezone import now


def view_vehicle(request):
    today = now().date()  # Get the current date
    data = Vehicle.objects.all()

    # Add a 'days_left' attribute for each field to use in the template
    for vehicle in data:
        vehicle.rc_days_left = (vehicle.rccardate - today).days if vehicle.rccardate else None
        vehicle.insurance_days_left = (vehicle.incurencedate - today).days if vehicle.incurencedate else None
        vehicle.permit_days_left = (vehicle.permitdate - today).days if vehicle.permitdate else None
        vehicle.tax_days_left = (vehicle.taxdate - today).days if vehicle.taxdate else None
        vehicle.emission_days_left = (vehicle.emissiondate - today).days if vehicle.emissiondate else None

    return render(request, 'view_vehicle.html', {'data': data})


def vehicle_edit(request, pk):
    data = Vehicle.objects.filter(id=pk).first()  # Retrieve a single object or None

    if request.method == "POST":
        vehicle_number = request.POST.get('vehicle_number')
        rcdate = request.POST.get('rcdate')
        incurencedate = request.POST.get('incurencedate')
        permitdate = request.POST.get('permitdate')
        taxdate = request.POST.get('taxdate')
        emissiondate = request.POST.get('emissiondate')

        data.vehicle_number = vehicle_number
        data.rccardate = rcdate
        data.incurencedate = incurencedate
        data.permitdate = permitdate
        data.taxdate = taxdate
        data.emissiondate = emissiondate

        data.save()

        # Redirect to a different URL after successful update
        base_url = reverse('view_vehicle')
        return redirect(base_url)

    return render(request, 'vehicle_edit.html', {'data': data})


def vehicle_delete(request, pk):
    udata = Vehicle.objects.get(id=pk)
    udata.delete()
    base_url = reverse('view_vehicle')
    return redirect(base_url)


def get_account_name(request):
    query = request.GET.get('query', '')
    if query:
        sender_names = Account.objects.filter(sender_name__icontains=query).values_list('sender_name',
                                                                                        flat=True).distinct()
        print('sender_names numbers:', list(sender_names))  # Debugging: check the data in the terminal
        return JsonResponse(list(sender_names), safe=False)
    return JsonResponse([], safe=False)


def get_consignor_name(request):
    query = request.GET.get('query', '')
    if query:
        sender_names = Consignor.objects.filter(sender_name__icontains=query).values_list('sender_name', flat=True)
        print('sender_names numbers:', list(sender_names))  # Debugging: check the data in the terminal
        return JsonResponse(list(sender_names), safe=False)
    return JsonResponse([], safe=False)


def get_sender_details(request):
    name = request.GET.get('name', '')
    if name:
        consignor = Consignor.objects.filter(sender_name=name).first()
        if consignor:
            data = {
                'sender_mobile': consignor.sender_mobile,
                'sender_email': consignor.sender_email,
                'sender_GST': consignor.sender_GST,
                'sender_address': consignor.sender_address,
                'sender_company': consignor.sender_company,
                'cust_id': consignor.cust_id,
            }
        else:
            data = {}
    else:
        data = {}

    return JsonResponse(data)


def get_consignee_name(request):
    query = request.GET.get('query', '')
    if query:
        receiver_names = Consignee.objects.filter(receiver_name__icontains=query).values_list('receiver_name',
                                                                                              flat=True)
        print('sender_names numbers:', list(receiver_names))  # Debugging: check the data in the terminal
        return JsonResponse(list(receiver_names), safe=False)
    return JsonResponse([], safe=False)


def get_rec_details(request):
    name = request.GET.get('name', '')
    if name:
        consignee = Consignee.objects.filter(receiver_name=name).first()
        if consignee:
            data = {
                'receiver_mobile': consignee.receiver_mobile,
                'receiver_GST': consignee.receiver_GST,
                'receiver_email': consignee.receiver_email,
                'receiver_address': consignee.receiver_address,
                'receiver_company': consignee.receiver_company,
                'cust_id': consignee.cust_id,
            }
        else:
            data = {}
    else:
        data = {}

    return JsonResponse(data)


def branchConsignment(request):
    if request.method == "POST":
        now = datetime.now().replace(microsecond=0)

        con_date = now.strftime("%Y-%m-%d")
        current_time = now.strftime("%H:%M:%S")

        uid = request.session.get('username')
        branch = Branch.objects.get(email=uid)
        branchemail = branch.email
        uname = branch.companyname
        username = branch.headname

        # Get the last track_id and increment it
        last_track_id = AddConsignment.objects.aggregate(Max('track_id'))['track_id__max']
        track_id = int(last_track_id) + 1 if last_track_id else 1000
        con_id = str(track_id)

        # Get the last Consignment_id and increment it
        last_con_id = AddConsignment.objects.aggregate(Max('Consignment_id'))['Consignment_id__max']
        Consignment_id = last_con_id + 1 if last_con_id else 1000
        Consignment_id = str(Consignment_id)

        last_cust_id = Consignee.objects.aggregate(Max('cust_id'))['cust_id__max']
        cust_id = last_cust_id + 1 if last_cust_id else 1
        cust_id = str(cust_id)

        last_consignor_id = Consignor.objects.aggregate(Max('cust_id'))['cust_id__max']
        consignor_id = last_consignor_id + 1 if last_consignor_id else 1000
        consignor_id = str(consignor_id)

        # Sender details
        send_name = request.POST.get('a1')
        send_mobile = request.POST.get('a2')
        send_address = request.POST.get('a4')
        sender_GST = request.POST.get('sendergst')

        # Receiver details
        rec_name = request.POST.get('a5')
        rec_mobile = request.POST.get('a6')
        rec_address = request.POST.get('a8')
        rec_GST = request.POST.get('receivergst')

        # Check if the Consignor already exists
        consignor, created = Consignor.objects.update_or_create(
            sender_name=send_name,
            defaults={
                'sender_mobile': send_mobile,
                'sender_address': send_address,
                'sender_GST': sender_GST,
                'branch': uname,
                'cust_id': consignor_id  # Temporarily set cust_id to None
            }
        )

        # If the consignor was created, assign a new cust_id
        if created:
            consignor_id = consignor.cust_id  # Use the newly created cust_id
        else:
            consignor_id = consignor.cust_id  # Retain existing cust_id if not created

        # Save consignor mobile number in UserLogin table
        UserLogin.objects.update_or_create(
            email=send_mobile,  # Use sender_mobile as the unique key
            defaults={
                'username': send_name,  # Set username to sender_name
                # Add any other fields as necessary
            }
        )

        # Check if the Consignee already exists
        consignee, created = Consignee.objects.update_or_create(
            receiver_name=rec_name,
            defaults={
                'receiver_mobile': rec_mobile,
                'receiver_address': rec_address,
                'receiver_GST': rec_GST,
                'branch': uname,
                'cust_id': cust_id  # Temporarily set cust_id to None
            }
        )

        # Save consignee mobile number in UserLogin table
        UserLogin.objects.update_or_create(
            email=rec_mobile,  # Use receiver_mobile as the unique key
            defaults={
                'username': rec_name,  # Set username to receiver_name
                # Add any other fields as necessary
            }
        )

        # If the consignee was created, assign a new cust_id
        if created:
            cust_id = consignee.cust_id  # Use the newly created cust_id
        else:
            cust_id = consignee.cust_id  # Retain existing cust_id if not created

        # Handling copies
        copies = []
        if request.POST.get('consignor_copy'):
            copies.append('Consignor Copy')
        if request.POST.get('consignee_copy'):
            copies.append('Consignee Copy')
        if request.POST.get('lorry_copy'):
            copies.append('Lorry Copy')
        copy_type = ', '.join(copies)  # Combine into a single string

        # Handling product entries
        products = request.POST.getlist('product[]')
        pieces = request.POST.getlist('pieces[]')

        # Other consignment details
        delivery = request.POST.get('delivery_option')
        prod_invoice = request.POST.get('prod_invoice')
        prod_price = request.POST.get('prod_price')
        weight = float(request.POST.get('weight') or 0)
        weightAmt = float(request.POST.get('weightAmt') or 0)
        freight = float(request.POST.get('freight') or 0)
        hamali = float(request.POST.get('hamali') or 0)
        door_charge = float(request.POST.get('door_charge') or 0)
        st_charge = float(request.POST.get('st_charge') or 0)
        cost = float(request.POST.get('cost') or 0)
        bal = float(request.POST.get('balance') or 0)
        pay_status = request.POST.get('payment')
        route_from = request.POST.get('from')
        route_to = request.POST.get('to')
        eway_bill = request.POST.get('ewaybill_no')

        utype = request.session.get('utype')
        branch_value = 'admin' if utype == 'admin' else uname

        # Determine the appropriate name based on pay_status
        if pay_status == 'Consigner_AC':
            account_name = send_name

        elif pay_status == 'Consignee_AC':
            account_name = rec_name
        else:
            account_name = send_name  # Default to sender_name if pay_status is neither
        # Create consignment records
        for product, piece in zip(products, pieces):
            if not product or not piece:
                continue
            AddConsignment.objects.create(
                track_id=con_id,
                Consignment_id=Consignment_id,
                sender_name=send_name,
                sender_mobile=send_mobile,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_mobile=rec_mobile,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                desc_product=product,
                pieces=piece,
                prod_invoice=prod_invoice,
                prod_price=prod_price,
                weightAmt=weightAmt,
                weight=weight,
                balance=bal,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total_cost=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                time=current_time,
                copy_type=copy_type,
                delivery=delivery,
                eway_bill=eway_bill,
                consignment_status='Pending',
                branchemail=branchemail,
                status='Consignment Added'

            )
            AddConsignmentTemp.objects.create(
                track_id=con_id,
                Consignment_id=Consignment_id,
                sender_name=send_name,
                sender_mobile=send_mobile,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_mobile=rec_mobile,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                desc_product=product,
                pieces=piece,
                prod_invoice=prod_invoice,
                prod_price=prod_price,
                weightAmt=weightAmt,
                weight=weight,
                balance=bal,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total_cost=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                time=current_time,
                copy_type=copy_type,
                delivery=delivery,
                eway_bill=eway_bill
            )
            Collection.objects.create(
                lrNo=con_id,
                sender_name=send_name,
                sender_address=send_address,
                sender_GST=sender_GST,
                receiver_name=rec_name,
                receiver_address=rec_address,
                receiver_GST=rec_GST,
                weightRate=weightAmt,
                weight=weight,
                balance=cost,
                freight=freight,
                hamali=hamali,
                door_charge=door_charge,
                st_charge=st_charge,
                route_from=route_from,
                route_to=route_to,
                total=cost,
                date=con_date,
                pay_status=pay_status,
                branch=branch_value,
                name=username,
                consignment_status='Pending',
                amount=0,

            )

        # Only handle the Account model if pay_status is 'Consigner_AC' or 'Consignee_AC'
        if pay_status in ['Consigner_AC', 'Consignee_AC']:
            try:
                # Confirm this block executes
                print(f"Attempting to save to Account with track_id {con_id}, sender_name {account_name}, cost {cost}")

                previous_balance_entry = Account.objects.filter(sender_name=account_name).order_by('-Date').first()
                previous_balance = float(previous_balance_entry.Balance) if previous_balance_entry else 0.0
                updated_balance = previous_balance + cost

                # Proceed to save/update account data
                Account.objects.update_or_create(
                    track_number=con_id,
                    defaults={
                        'sender_name': account_name,
                        'Date': now,
                        'particulars': f'LR: {con_id}',
                        'debit': cost,
                        'credit': 0,
                        'TrType': 'sal',
                        'Balance': updated_balance,
                        'Branch': branch_value,  # Use Branch instead of branch
                        'headname': username
                    }
                )

                print(f"Account record for track_id {con_id} saved successfully.")
            except Exception as e:
                print(f"Error updating account: {str(e)}")

        return redirect('branchprintConsignment', track_id=con_id)
    return render(request, 'branchConsignment.html')


from django.utils.safestring import mark_safe
from django.core.exceptions import ObjectDoesNotExist

def branchprintConsignment(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity

    try:
        # Filter consignments by track_id
        consignments = AddConsignment.objects.filter(track_id=track_id)
        uid = request.session.get('username')
        branchdetails = Branch.objects.get(email=uid)

        if not consignments.exists():
            return render(request, '404.html')  # Handle case where no consignments are found.

        # Preprocess the `services` field to replace commas with `<br/>`
        branchdetails.services = mark_safe(branchdetails.services.replace(',', '<br/>'))

        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'desc_product': consignment.desc_product,
            }
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)
            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    return render(request, 'branchprintConsignment.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),  # Include the aggregated copy types
        'totalqty': totalqty  # Pass total quantity to the template
    })


from django.shortcuts import render
from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date

def branchviewconsignment(request):
    uid = request.session.get('username')
    grouped_userdata = {}

    if uid:
        try:
            # Fetch branch details for the logged-in user
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname  # Adjust if the branch info is stored differently

            # Get filters from POST request
            consigner = request.POST.get('consigner')
            consigee = request.POST.get('consignee')
            track_id = request.POST.get('lrno')
            from_date_str = request.POST.get('from_date')
            to_date_str = request.POST.get('to_date')

            # Parse date strings to date objects
            from_date = parse_date(from_date_str) if from_date_str else None
            to_date = parse_date(to_date_str) if to_date_str else None

            # Fetch consignments for the branch
            consignments = AddConsignment.objects.filter(branch=user_branch)

            if consigner:
                consignments = consignments.filter(sender_name=consigner)
            if consigee:
                consignments = consignments.filter(receiver_name=consigee)
            if track_id:
                consignments = consignments.filter(track_id=track_id)

            if from_date and to_date:
                consignments = consignments.filter(date__range=(from_date, to_date))
            elif from_date:
                consignments = consignments.filter(date__gte=from_date)
            elif to_date:
                consignments = consignments.filter(date__lte=to_date)

            # Fetch LRno values from Tripsheetprem table
            tripsheet_lrnos = set(TripSheetPrem.objects.values_list('LRno', flat=True))

            # Group consignments by track_id and check for matches
            for consignment in consignments:
                track_id = consignment.track_id
                if track_id not in grouped_userdata:
                    grouped_userdata[track_id] = {
                        'route_from': consignment.route_from,
                        'route_to': consignment.route_to,
                        'sender_name': consignment.sender_name,
                        'sender_mobile': consignment.sender_mobile,
                        'receiver_name': consignment.receiver_name,
                        'receiver_mobile': consignment.receiver_mobile,
                        'total_cost': consignment.total_cost,
                        'pieces': 0,
                        'weight': consignment.weight,
                        'pay_status': consignment.pay_status,
                        'products': [],
                        'match_found': track_id in tripsheet_lrnos  # Check if track_id matches LRno
                    }
                # Aggregate total pieces
                grouped_userdata[track_id]['pieces'] += consignment.pieces
                # Concatenate product descriptions
                product_detail = consignment.desc_product
                grouped_userdata[track_id]['products'].append(product_detail)

        except ObjectDoesNotExist:
            pass

    # Convert product details list to a single string
    for track_id, details in grouped_userdata.items():
        details['products'] = ', '.join(details['products'])

    # Render the template with grouped data
    return render(request, 'branchviewConsignment.html', {'grouped_userdata': grouped_userdata})


def branchMaster(request):
    uid = request.session['username']
    email = Branch.objects.get(email=uid)
    bid = email.id
    data = Branch.objects.filter(id=bid).first()  # Retrieve a single object or None
    if request.method == "POST":
        companyname = request.POST.get('companyname')
        phonenumber = request.POST.get('phonenumber')
        email = request.POST.get('email')
        gst = request.POST.get('gst')
        address = request.POST.get('address')
        image = request.POST.get('image')

        # Update the object
        data.companyname = companyname
        data.phonenumber = phonenumber
        data.email = email
        data.gst = gst
        data.address = address
        data.image = image

        data.save()

        # Redirect to a different URL after successful update
        base_url = reverse('branchMaster')
        return redirect(base_url)

    return render(request, 'branchMaster.html', {'data': data})


def branchconsignment_edit(request, pk):
    # The main consignment object, your container
    userdata = get_object_or_404(AddConsignment, track_id=pk)

    # Fetch all products under this consignment
    products = AddConsignment.objects.filter(track_id=pk)

    # Fetch consignment temp data if it exists
    userdatatemp = AddConsignmentTemp.objects.filter(track_id=pk).first()

    if request.method == "POST":
        userdata.route_from = request.POST.get('from')
        userdata.route_to = request.POST.get('to')

        userdata.sender_name = request.POST.get('a1')
        userdata.receiver_name = request.POST.get('a5')
        userdata.total_cost = request.POST.get('cost')
        userdata.weight = request.POST.get('weight')
        userdata.balance = request.POST.get('bal')
        userdata.pay_status = request.POST.get('payment')

        # Add any additional fields you want to update here
        userdata.sender_address = request.POST.get('a4')
        userdata.sender_mobile = request.POST.get('a2')
        userdata.sender_GST = request.POST.get('sendergst')
        userdata.receiver_address = request.POST.get('a8')
        userdata.receiver_mobile = request.POST.get('a6')
        userdata.receiver_GST = request.POST.get('receivergst')
        userdata.prod_invoice = request.POST.get('prod_invoice')
        userdata.prod_price = request.POST.get('prod_price')
        userdata.weightAmt = request.POST.get('weightAmt')
        userdata.freight = request.POST.get('freight')
        userdata.hamali = request.POST.get('hamali')
        userdata.door_charge = request.POST.get('door_charge')
        userdata.st_charge = request.POST.get('st_charge')
        userdata.save()

        # Update AddConsignmentTemp only if userdatatemp exists
        if userdatatemp:
            userdata.route_from = request.POST.get('from')
            userdata.route_to = request.POST.get('to')
            userdatatemp.sender_name = request.POST.get('a1')
            userdatatemp.receiver_name = request.POST.get('a5')
            userdatatemp.total_cost = request.POST.get('cost')
            userdatatemp.weight = request.POST.get('weight')
            userdatatemp.balance = request.POST.get('bal')
            userdatatemp.pay_status = request.POST.get('payment')

            # Add any additional fields you want to update here
            userdatatemp.sender_address = request.POST.get('a4')
            userdatatemp.sender_mobile = request.POST.get('a2')
            userdatatemp.sender_GST = request.POST.get('sendergst')
            userdatatemp.receiver_address = request.POST.get('a8')
            userdatatemp.receiver_mobile = request.POST.get('a6')
            userdatatemp.receiver_GST = request.POST.get('receivergst')
            userdatatemp.prod_invoice = request.POST.get('prod_invoice')
            userdatatemp.prod_price = request.POST.get('prod_price')
            userdatatemp.weightAmt = request.POST.get('weightAmt')
            userdatatemp.freight = request.POST.get('freight')
            userdatatemp.hamali = request.POST.get('hamali')
            userdatatemp.door_charge = request.POST.get('door_charge')
            userdatatemp.st_charge = request.POST.get('st_charge')
            userdatatemp.save()

        # Handling products - whether updating or adding new ones
        product_ids = request.POST.getlist('product_id[]')
        products_list = request.POST.getlist('product[]')
        pieces_list = request.POST.getlist('pieces[]')

        # Updating products in AddConsignment
        for product_id, product_desc, piece in zip(product_ids, products_list, pieces_list):
            if product_desc and piece:  # If we have product data
                if product_id:  # Existing product - update
                    product_obj = AddConsignment.objects.get(id=product_id)
                    product_obj.desc_product = product_desc
                    product_obj.pieces = piece
                    product_obj.save()
                else:  # New product - add to consignment
                    new_product = AddConsignment(
                        track_id=userdata.track_id,  # Use the same track_id for all products
                        desc_product=product_desc,
                        pieces=piece,
                        sender_name=userdata.sender_name,  # Ensuring consistency
                        receiver_name=userdata.receiver_name
                    )
                    new_product.save()

        # Handling products in AddConsignmentTemp if userdatatemp exists
        if userdatatemp:
            product_idst = request.POST.getlist('product_id[]')
            products_listt = request.POST.getlist('product[]')
            pieces_listt = request.POST.getlist('pieces[]')

            for product_idt, product_desct, piecet in zip(product_idst, products_listt, pieces_listt):
                if product_desct and piecet:  # Ensure we have valid data
                    try:
                        # Try to get the existing product in AddConsignmentTemp
                        product_obj = AddConsignmentTemp.objects.get(id=product_idt)
                        product_obj.desc_product = product_desct
                        product_obj.pieces = piecet
                        product_obj.save()
                    except ObjectDoesNotExist:
                        # If product doesn't exist, create a new one
                        new_product = AddConsignmentTemp(
                            track_id=userdata.track_id,  # Use the same track_id for all products
                            desc_product=product_desct,
                            pieces=piecet,
                            sender_name=userdata.sender_name,  # Ensuring consistency
                            receiver_name=userdata.receiver_name
                        )
                        new_product.save()

        # After saving everything, redirect to view the consignment page
        return redirect(reverse('branchviewconsignment'))

    return render(request, 'branchconsignment_edit.html', {'userdata': userdata, 'products': products})


def branchconsignment_delete(request, pk):
    udata = AddConsignment.objects.get(id=pk)
    udata.delete()
    base_url = reverse('view_consignment')
    return redirect(base_url)

def branchinvoiceConsignment(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity

    try:
        # Filter consignments by track_id
        consignments = AddConsignment.objects.filter(track_id=track_id)
        uid = request.session.get('username')
        branchdetails = Branch.objects.get(email=uid)

        if not consignments.exists():
            return render(request, '404.html')  # Handle case where no consignments are found.

        # Preprocess the `services` field to replace commas with `<br/>`
        branchdetails.services = mark_safe(branchdetails.services.replace(',', '<br/>'))

        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'desc_product': consignment.desc_product,
            }
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)
            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    return render(request, 'branchinvoiceConsignment.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),  # Include the aggregated copy types
        'totalqty': totalqty  # Pass total quantity to the template
    })




def branchaddTrack(request):
    userid = request.session.get('username')
    userdata = Branch.objects.get(email=userid)
    uname = userdata.companyname
    consignments = AddConsignment.objects.filter(branch=uname).order_by('-id')

    if request.method == "POST":
        now = datetime.datetime.now()
        con_date = now.strftime("%Y-%m-%d")

        track_id = request.POST.get('a1')
        status = request.POST.get('status')  # Retrieve status from the form

        # Retrieve custom status if "Other" is selected
        if status == "Other":
            custom_status = request.POST.get('a2')
        else:
            custom_status = None

        # Retrieve username from session and fetch the corresponding branch
        uid = request.session.get('username')

        if uid:
            userdata = Branch.objects.get(email=uid)
            uname = userdata.companyname

            # Check utype to determine the branch value
            utype = request.session.get('utype')
            branch_value = 'admin' if utype == 'admin' else uname

            # Filter consignment data based on the branch
            consignments = AddConsignment.objects.filter(branch=uname).order_by('-id')

            # Create AddTrack object
            AddTrack.objects.create(
                track_id=track_id,
                description=status,
                date=con_date,
                branch=branch_value
            )

        else:
            # Handle the case where session data is missing
            consignments = AddConsignment.objects.none()
            return render(request, 'branchaddTrack.html', {'consignments': consignments, 'msg': 'Session data missing'})

    return render(request, 'branchaddTrack.html', {'consignments': consignments})


def branchsearch_results(request):
    tracker_id = request.GET.get('tracker_id')
    userid = request.session.get('username')
    userdata = Branch.objects.get(email=userid)
    uname = userdata.companyname
    consignments = AddConsignment.objects.filter(branch=uname).order_by('-id')

    if tracker_id:
        try:
            trackers = AddTrack.objects.filter(track_id=tracker_id)
            if trackers.exists():
                return render(request, 'branchsearch_results.html',
                              {'trackers': trackers, 'consignments': consignments})
            else:
                message = f"No tracking information found for ID: {tracker_id}"
                return render(request, 'branchsearch_results.html', {'message': message, 'consignments': consignments})
        except Exception as e:
            message = f"Error occurred: {str(e)}"
            return render(request, 'branchsearch_results.html', {'message': message, 'consignments': consignments})
    else:
        return render(request, 'branchsearch_results.html',
                      {'message': "Please enter a tracker ID.", 'consignments': consignments})


def branchtrack_delete(request, pk):
    udata = AddTrack.objects.get(id=pk)
    udata.delete()
    base_url = reverse('branchsearch_results')
    return redirect(base_url)


def get_vehicle_numbers(request):
    query = request.GET.get('query', '')
    if query:
        vehicle_numbers = Vehicle.objects.filter(vehicle_number__icontains=query).values_list('vehicle_number',
                                                                                              flat=True)
        print('Vehicle numbers:', list(vehicle_numbers))  # Debugging: check the data in the terminal
        return JsonResponse(list(vehicle_numbers), safe=False)
    return JsonResponse([], safe=False)


def get_driver_name(request):
    query = request.GET.get('query', '')
    if query:
        driver_name = Driver.objects.filter(driver_name__icontains=query).values_list('driver_name', flat=True)
        print('Driver Name:', list(driver_name))  # Debugging: check the data in the terminal
        return JsonResponse(list(driver_name), safe=False)
    return JsonResponse([], safe=False)


def get_branch(request):
    query = request.GET.get('query', '')
    if query:
        companyname = Branch.objects.filter(companyname__icontains=query).values_list('companyname', flat=True)
        print('Branch Name:', list(companyname))  # Debugging: check the data in the terminal
        return JsonResponse(list(companyname), safe=False)
    return JsonResponse([], safe=False)


def get_destination(request):
    query = request.GET.get('query', '')
    if query:
        # Filter and get distinct route_to values
        route_to = AddConsignment.objects.filter(route_to__icontains=query).values_list('route_to',
                                                                                        flat=True).distinct()
        print('Distinct route_to numbers:', list(route_to))  # Debugging: check the data in the terminal
        return JsonResponse(list(route_to), safe=False)
    return JsonResponse([], safe=False)


from collections import defaultdict



def addTripSheet(request):
    route_to = AddConsignmentTemp.objects.values_list('route_to', flat=True).distinct()
    addtrip = defaultdict(
        lambda: {'desc_product': [], 'pieces': 0, 'receiver_name': '','receiver_address': '', 'receiver_GST': '','sender_name': '','sender_GST': '','sender_address': '','pay_status': '', 'route_to': '', 'total': '',
                 'weightAmt': '', 'freight': '', 'hamali': '', 'door_charge': '', 'st_charge': ''}
    )

    no_data_found = False  # Flag to check if data was found

    try:
        # Fetch all consignment data without any filtering
        consignments = AddConsignmentTemp.objects.all()

        if consignments.exists():
            for consignment in consignments:
                consignment_data = addtrip[consignment.track_id]
                consignment_data['desc_product'].append(consignment.desc_product)
                consignment_data['pieces'] += consignment.pieces
                consignment_data['route_to'] = consignment.route_to
                consignment_data['receiver_name'] = consignment.receiver_name
                consignment_data['receiver_GST'] = consignment.receiver_GST
                consignment_data['receiver_address'] = consignment.receiver_address
                consignment_data['sender_name'] = consignment.sender_name
                consignment_data['sender_GST'] = consignment.sender_GST
                consignment_data['sender_address'] = consignment.sender_address
                consignment_data['pay_status'] = consignment.pay_status
                consignment_data['total_cost'] = consignment.total_cost
                consignment_data['weightAmt'] = consignment.weightAmt
                consignment_data['freight'] = consignment.freight
                consignment_data['hamali'] = consignment.hamali
                consignment_data['door_charge'] = consignment.door_charge
                consignment_data['st_charge'] = consignment.st_charge
        else:
            no_data_found = True  # Set the flag if no data is found

        # Convert the defaultdict to a list of dictionaries
        addtrip = [
            {
                'track_id': track_id,
                'desc_product': ', '.join(consignment_data['desc_product']),
                'pieces': consignment_data['pieces'],
                'route_to': consignment_data['route_to'],
                'receiver_name': consignment_data['receiver_name'],
                'receiver_GST': consignment_data['receiver_GST'],
                'receiver_address': consignment_data['receiver_address'],
                'sender_name': consignment_data['sender_name'],
                'sender_GST': consignment_data['sender_GST'],
                'sender_address': consignment_data['sender_address'],
                'pay_status': consignment_data['pay_status'],
                'total_cost': consignment_data['total_cost'],
                'weightAmt': consignment_data['weightAmt'],
                'freight': consignment_data['freight'],
                'hamali': consignment_data['hamali'],
                'door_charge': consignment_data['door_charge'],
                'st_charge': consignment_data['st_charge']
            }
            for track_id, consignment_data in addtrip.items()
        ]

    except Exception as e:
        addtrip = []
        no_data_found = True  # Set the flag if an error occurs

    return render(request, 'addTripSheet.html', {
        'route_to': route_to,
        'trip': addtrip,
        'no_data_found': no_data_found  # Pass the flag to the template
    })


def saveTripSheetList(request):
    print("saveTripSheet function called")
    if request.method == 'POST':
        print("POST request received")  # Debugging statement

        uid = request.session.get('username')
        if uid:
            try:
                branch = Branch.objects.get(email=uid)
                branchname = branch.companyname
                username = branch.headname

                now = datetime.now()
                con_date = now.strftime("%Y-%m-%d")
                current_time = now.strftime("%H:%M:%S")

                total_rows = int(request.POST.get('total_rows', 0))

                selected_rows = request.POST.getlist('selected_rows')

                for i in range(1, total_rows + 1):
                    if str(i) in selected_rows:  # Only process if the row is selected
                        track_id = request.POST.get(f'track_id_{i}')
                        pieces = request.POST.get(f'pieces_{i}')
                        desc_product = request.POST.get(f'desc_product_{i}')
                        route_to = request.POST.get(f'route_to_{i}')
                        receiver_name = request.POST.get(f'receiver_name_{i}')
                        receiver_GST = request.POST.get(f'receiver_GST_{i}')
                        receiver_address = request.POST.get(f'receiver_address_{i}')
                        sender_name = request.POST.get(f'sender_name_{i}')
                        sender_GST = request.POST.get(f'sender_GST_{i}')
                        sender_address = request.POST.get(f'sender_address_{i}')
                        pay_status = request.POST.get(f'pay_status_{i}')
                        total_cost = request.POST.get(f'total_cost{i}')
                        weightAmt = request.POST.get(f'weightAmt{i}')
                        freight = request.POST.get(f'freight{i}')
                        hamali = request.POST.get(f'hamali{i}')
                        door_charge = request.POST.get(f'door_charge{i}')
                        st_charge = request.POST.get(f'st_charge{i}')

                        print(
                            f"Track ID: {track_id}, Pieces: {pieces}, Description: {desc_product}, Route: {route_to}, Receiver: {receiver_name}, Pay Status: {pay_status}, total_cost:{total_cost},weightAmt:{weightAmt},freight:{freight},hamali:{hamali},door_charge:{door_charge},st_charge:{st_charge}")  # Debugging statement

                        # Save to TripSheetTemp
                        TripSheetTemp.objects.create(
                            LRno=track_id,
                            qty=pieces,
                            desc=desc_product,
                            dest=route_to,
                            consignee=receiver_name,
                            consignor=sender_name,
                            consignorGST=sender_GST,
                            consignorAddress=sender_address,
                            consigneeGST=receiver_GST,
                            consigneeAddress=receiver_address,
                            pay_status=pay_status,
                            branch=branchname,
                            username=username,
                            Date=con_date,
                            total_cost=total_cost,
                            weightAmt=weightAmt,
                            freight=freight,
                            hamali=hamali,
                            door_charge=door_charge,
                            st_charge=st_charge,
                        )

                        # Delete from AddConsignmentTemp
                        AddConsignmentTemp.objects.filter(track_id=track_id).delete()

                        print(
                            f"Data for Track ID {track_id} saved and deleted from AddConsignmentTemp successfully.")  # Debugging statement
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
        else:
            print("No username found in session.")  # Debugging statement

        return redirect('addTripSheet')  # Replace with your desired success URL

    print("Not a POST request, redirecting back to form.")  # Debugging statement
    return render(request, 'addTripSheet.html')  # Redirect back to the form if not a POST request


def addTripSheetList(request):
    addtrip = []  # Initialize an empty list to store trip details
    uid = request.session.get('username')
    no_data_found = False  # Flag to check if no data is found

    if uid:
        try:
            # Fetch the user's branch from the session
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            if request.method == 'POST':
                # Get the selected date from the form
                date = request.POST.get('date')

                if date:
                    # Query TripSheetTemp table based on the selected date and user's branch
                    consignments = TripSheetTemp.objects.filter(
                        Date=date,
                        branch=user_branch
                    )

                    # Check if consignments exist
                    if consignments.exists():
                        # Iterate through the results and prepare the data for the template
                        addtrip = [
                            {
                                'track_id': consignment.LRno,
                                'desc': consignment.desc,
                                'qty': consignment.qty,
                                'dest': consignment.dest,
                                'consignee': consignment.consignee,
                                'consigneeGST': consignment.consigneeGST,
                                'consigneeAddress': consignment.consigneeAddress,
                                'consignor': consignment.consignor,
                                'consignorGST': consignment.consignorGST,
                                'consignorAddress': consignment.consignorAddress,
                                'pay_status': consignment.pay_status,
                                'total_cost': consignment.total_cost,
                                'weightAmt': consignment.weightAmt,
                                'freight': consignment.freight,
                                'hamali': consignment.hamali,
                                'door_charge': consignment.door_charge,
                                'st_charge': consignment.st_charge
                            }
                            for consignment in consignments
                        ]
                    else:
                        no_data_found = True  # Set the flag if no data is found

        except Branch.DoesNotExist:
            addtrip = []
            no_data_found = True  # Set the flag if the branch does not exist

    # Render the template with the trip data and no_data_found flag
    return render(request, 'addTripSheetList.html', {
        'trip': addtrip,
        'no_data_found': no_data_found,
    })


def saveTripSheet(request):
    print("saveTripSheet function called")

    if request.method == 'POST':
        print("POST request received")  # Debugging statement

        # Generate trip_id
        last_trip_id = TripSheetPrem.objects.aggregate(Max('trip_id'))['trip_id__max']
        trip_id = int(last_trip_id) + 1 if last_trip_id else 1000  # Start from a defined base if no entries exist
        con_id = str(trip_id)

        uid = request.session.get('username')
        if uid:
            try:
                branch = Branch.objects.get(email=uid)
                branchname = branch.companyname
                username = branch.headname

                now = datetime.now()
                con_date = now.strftime("%Y-%m-%d")
                current_time = now.strftime("%H:%M:%S")

                # Get form data
                vehicle = request.POST.get('vehical')
                drivername = request.POST.get('drivername')
                adv = request.POST.get('advance')
                ltrate = request.POST.get('ltrate') or 0
                ltr = request.POST.get('liter') or 0

                literate = float(ltrate)
                liter = float(ltr)
                diesel_total = literate * liter


                # Save to Disel table
                Disel.objects.create(
                    Date=con_date,
                    vehicalno=vehicle,
                    drivername=drivername,
                    ltrate=ltrate,
                    liter=ltr,
                    total=diesel_total,  # Diesel total cost
                    trip_id=con_id
                )

                total_rows = int(request.POST.get('total_rows', 0))

                print(f"Vehicle: {vehicle}, Driver Name: {drivername}")  # Debugging statement

                for i in range(1, total_rows + 1):
                    track_id = request.POST.get(f'track_id_{i}')
                    desc = request.POST.get(f'desc_{i}')
                    qty = request.POST.get(f'qty_{i}')
                    dest = request.POST.get(f'dest_{i}')
                    consignee = request.POST.get(f'consignee_{i}')
                    consigneeGST = request.POST.get(f'consigneeGST_{i}')
                    consigneeAddress = request.POST.get(f'consigneeAddress_{i}')
                    consignor = request.POST.get(f'consignor_{i}')
                    consignorGST = request.POST.get(f'consignorGST_{i}')
                    consignorAddress = request.POST.get(f'consignorAddress_{i}')
                    total_cost = request.POST.get(f'total_cost_{i}')
                    pay_status = request.POST.get(f'pay_status_{i}')
                    weightAmt = request.POST.get(f'weightAmt_{i}')
                    freight = request.POST.get(f'freight_{i}')
                    hamali = request.POST.get(f'hamali_{i}')
                    door_charge = request.POST.get(f'door_charge_{i}')
                    st_charge = request.POST.get(f'st_charge_{i}')

                    print(
                        f"Track ID: {track_id}, Description: {desc}, Quantity: {qty}, Route: {dest}, Receiver: {consignee}")  # Debugging

                    # Save to TripSheetPrem
                    TripSheetPrem.objects.create(
                        LRno=track_id,
                        qty=qty,
                        desc=desc,
                        dest=dest,
                        consignee=consignee,
                        consigneeGST=consigneeGST,
                        consigneeAddress=consigneeAddress,
                        consignor=consignor,
                        consignorGST=consignorGST,
                        consignorAddress=consignorAddress,
                        pay_status=pay_status,
                        VehicalNo=vehicle,
                        DriverName=drivername,
                        branch=branchname,
                        username=username,
                        Date=con_date,
                        Time=current_time,
                        AdvGiven=adv,
                        LTRate=ltrate,
                        Ltr=ltr,
                        total_cost=total_cost,
                        weightAmt=float(weightAmt),
                        freight=freight,
                        hamali=hamali,
                        door_charge=door_charge,
                        st_charge=st_charge,
                        trip_id=con_id,
                        status='TripSheet Added',
                    )

                    # Delete from AddConsignmentTemp
                    TripSheetTemp.objects.filter(LRno=track_id).delete()

                    print(f"Data for Track ID {track_id} saved successfully.")  # Debugging statement
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
        else:
            print("No username found in session.")  # Debugging statement

        return redirect('addTripSheetList')  # Replace with your desired success URL

    return render(request, 'addTripSheetList.html')  # Redirect back if not a POST request


from django.db.models import Sum


def tripSheet(request):
    return render(request, 'tripSheet.html')


def tripSheetList(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            if request.method == 'POST':
                vehicle_number = request.POST.get('vehical')
                date = request.POST.get('t3')

                if date:
                    trips = TripSheetPrem.objects.filter(
                        VehicalNo=vehicle_number,
                        Date=date,
                        branch=user_branch
                    )
                    # Calculate total quantity
                    total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

                    # Aggregate data based on pay_status
                    statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
                    for status in statuses:
                        status_trips = trips.filter(pay_status=status)
                        summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                        summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                        summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                        summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                        summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                        summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                        # Update grand totals
                        grand_total[status] = summary[status]['total_cost']
                        grand_total['grand_freight'] += summary[status]['freight']
                        grand_total['grand_hamali'] += summary[status]['hamali']
                        grand_total['grand_st_charge'] += summary[status]['st_charge']
                        grand_total['grand_door_charge'] += summary[status]['door_charge']
                        grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                        grand_total['grand_total'] += summary[status]['total_cost']

                    # Calculate the total value using the first row
                    if trips.exists():
                        first_trip = trips.first()
                        total_ltr_value = float(
                            first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                        total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                        total_value = total_ltr_value + total_adv_given
                    else:
                        total_value = 0.0

        except ObjectDoesNotExist:
            trips = TripSheetTemp.objects.none()

    return render(request, 'TripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary
    })


@require_POST
def delete_trip_sheet_data(request):
    vehicle_number = request.POST.get('vehical')
    date = request.POST.get('t3')
    uid = request.session.get('username')

    print(f"Received vehicle_number: {vehicle_number}, date: {date}, uid: {uid}")

    if uid and vehicle_number and date:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname
            TripSheetTemp.objects.filter(
                VehicalNo=vehicle_number,
                Date=date,
                branch=user_branch
            ).delete()
            return JsonResponse({'status': 'success'})
        except ObjectDoesNotExist:
            print("Branch does not exist.")
            return JsonResponse({'status': 'error', 'message': 'Branch does not exist'})

    print("Invalid parameters received.")
    return JsonResponse({'status': 'error', 'message': 'Invalid parameters'})


def viewTripSheetList(request):
    grouped_trips = []
    uid = request.session.get('username')

    if uid:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            # Group by VehicalNo and Date, and annotate with count
            if request.method == 'POST' and request.POST.get('t3'):
                date = request.POST.get('t3')
                grouped_trips = (
                    TripSheetPrem.objects
                    .filter(Date=date, branch=user_branch)
                    .values('VehicalNo', 'Date')
                    .annotate(trip_count=Count('id'))
                )
            else:
                # Fetch all trips for the branch without date filtering
                grouped_trips = (
                    TripSheetPrem.objects
                    .filter(branch=user_branch)
                    .values('VehicalNo', 'Date')
                    .annotate(trip_count=Count('id'))
                )
        except ObjectDoesNotExist:
            grouped_trips = []

    return render(request, 'viewTripSheetList.html', {
        'grouped_trips': grouped_trips
    })

def editTripSheetList(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            if request.method == 'POST':
                vehicle_number = request.POST.get('vehical')
                date_str = request.POST.get('t3')

                if date_str:
                    # Directly use date_str if it's in yyyy-mm-dd format
                    date = date_str

                    # Filter trips based on the vehicle number, date, and branch
                    trips = TripSheetPrem.objects.filter(
                        VehicalNo=vehicle_number,
                        Date=date,
                        branch=user_branch
                    )
                    # Calculate total quantity
                    total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

                    # Aggregate data based on pay_status
                    statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
                    for status in statuses:
                        status_trips = trips.filter(pay_status=status)
                        summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                        summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                        summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                        summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                        summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                        summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                        # Update grand totals
                        grand_total[status] = summary[status]['total_cost']
                        grand_total['grand_freight'] += summary[status]['freight']
                        grand_total['grand_hamali'] += summary[status]['hamali']
                        grand_total['grand_st_charge'] += summary[status]['st_charge']
                        grand_total['grand_door_charge'] += summary[status]['door_charge']
                        grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                        grand_total['grand_total'] += summary[status]['total_cost']

                    # Calculate the total value using the first row
                    if trips.exists():
                        first_trip = trips.first()
                        total_ltr_value = float(
                            first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                        total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                        total_value = total_ltr_value + total_adv_given
                    else:
                        total_value = 0.0

        except Branch.DoesNotExist:
            trips = TripSheetTemp.objects.none()

    return render(request, 'editTripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary
    })


def update_view(request):
    if request.method == "POST":
        trip_id = request.POST.get("trip_id")
        print(f"Received trip_id: {trip_id}")  # Debugging line

        # Fetch all records with the matching trip_id
        trips = TripSheetPrem.objects.filter(trip_id=trip_id)

        if trips.exists():
            print(f"Found {trips.count()} trip records to update")
            for trip in trips:
                # Update the fields for each trip
                trip.LTRate = request.POST.get("ltrate")
                trip.Ltr = request.POST.get("ltr")
                trip.AdvGiven = request.POST.get("advgiven")
                trip.commission = request.POST.get("commission")
                trip.save()

            # Redirect after saving
            return redirect('viewTripSheetList')  # Replace with your success URL
        else:
            print("No trip records found")
            return render(request, 'editTripSheetList.html',
                          {'error_message': 'No trips found with the provided trip_id.'})

    return render(request, 'editTripSheetList.html')  # Replace with your template


def printTripSheetList(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            vehical_no = request.POST.get('vehical')
            date = request.POST.get('t3')

            # Filter trips based on VehicleNo, Date, and branch
            trips = TripSheetPrem.objects.filter(
                VehicalNo=vehical_no,
                Date=date,
                branch=user_branch
            )

            # Calculate total quantity
            total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

            # Aggregate data based on pay_status
            statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
            for status in statuses:
                status_trips = trips.filter(pay_status=status)
                summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                # Update grand totals
                grand_total[status] = summary[status]['total_cost']
                grand_total['grand_freight'] += summary[status]['freight']
                grand_total['grand_hamali'] += summary[status]['hamali']
                grand_total['grand_st_charge'] += summary[status]['st_charge']
                grand_total['grand_door_charge'] += summary[status]['door_charge']
                grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                grand_total['grand_total'] += summary[status]['total_cost']

            # Calculate the total value using the first row
            if trips.exists():
                first_trip = trips.first()
                total_ltr_value = float(
                    first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                total_value = total_ltr_value
            else:
                total_value = 0.0

        except Branch.DoesNotExist:
            trips = TripSheetPrem.objects.none()  # Handle case where Branch does not exist

    return render(request, 'printTripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary
    })

from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.http import HttpResponse

def exportTripSheetToExcel(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Branch.objects.get(email=uid)
            user_branch = branch.companyname

            vehical_no = request.GET.get('vehical')
            date = request.GET.get('t3')

            trips = TripSheetPrem.objects.filter(
                VehicalNo=vehical_no,
                Date=date,
                branch=user_branch
            )

            total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

            statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
            for status in statuses:
                status_trips = trips.filter(pay_status=status)
                summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                grand_total[status] = summary[status]['total_cost']
                grand_total['grand_freight'] += summary[status]['freight']
                grand_total['grand_hamali'] += summary[status]['hamali']
                grand_total['grand_st_charge'] += summary[status]['st_charge']
                grand_total['grand_door_charge'] += summary[status]['door_charge']
                grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                grand_total['grand_total'] += summary[status]['total_cost']

            if trips.exists():
                first_trip = trips.first()
                total_ltr_value = float(first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                total_value = total_ltr_value
            else:
                total_value = 0.0

        except Branch.DoesNotExist:
            trips = TripSheetPrem.objects.none()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trip Sheet"

    # Merging cells for Branch Name
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    sheet["A1"] = branch.companyname if uid else "N/A"
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Merging cells for Driver, Vehicle, and Date details
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    driver_name = first_trip.DriverName if trips.exists() else "N/A"
    sheet["A2"] = f"Driver: {driver_name}, VehicalNo: {vehical_no}, Date: {date}"
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers
    headers = [
        "SL/N", "LR No", "QTY", "DESC OF GOODS", "DEST", "CONSIGNOR", "CONSIGNOR GST",
        "CONSIGNEE", "CONSIGNEE GST", "TO PAY", "PAID", "CSER A/C", "CSGEE A/C"
    ]
    sheet.append(headers)

    # Trip Data
    for idx, trip in enumerate(trips, start=1):
        sheet.append([
            idx, trip.LRno, trip.qty, trip.desc, trip.dest, trip.consignor,
            trip.consignorGST, trip.consignee, trip.consigneeGST,
            trip.total_cost if trip.pay_status == "ToPay" else "",
            trip.total_cost if trip.pay_status == "Paid" else "",
            trip.total_cost if trip.pay_status == "Consigner_AC" else "",
            trip.total_cost if trip.pay_status == "Consignee_AC" else "",
        ])

    # Summary Table
    sheet.append([])
    sheet.append(["HEAD", "Weight Amt", "FREIGHT", "HAMALI", "ST CHARGE", "DOOR CHARGE", "TOTAL"])
    for status, details in summary.items():
        sheet.append([
            status, details['weightAmt'], details['freight'], details['hamali'],
            details['st_charge'], details['door_charge'], details['total_cost']
        ])

    # Grand Total
    sheet.append([])
    sheet.append([
        "Total", grand_total['grand_weightAmt'], grand_total['grand_freight'],
        grand_total['grand_hamali'], grand_total['grand_st_charge'],
        grand_total['grand_door_charge'], grand_total['grand_total']
    ])

    # Export File
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="trip_sheet.xlsx"'
    workbook.save(response)
    return response


def adminTripSheet(request):
    grouped_trips = []

    if request.method == 'POST':
        vehicle_number = request.POST.get('vehical')
        branch = request.POST.get('t2')
        date = request.POST.get('t3')

        # Apply filters only if at least one filter value is provided
        filters = {}
        if date:
            filters['Date'] = date
        if vehicle_number:
            filters['VehicalNo'] = vehicle_number
        if branch:
            filters['branch'] = branch

        # Group by VehicalNo and Date, and annotate with count
        grouped_trips = (
            TripSheetPrem.objects
            .filter(**filters)
            .values('VehicalNo', 'Date', 'branch')
            .annotate(trip_count=Count('id'))
        )
    else:
        # When no POST request, display all grouped trips
        grouped_trips = (
            TripSheetPrem.objects
            .values('VehicalNo', 'Date', 'branch')
            .annotate(trip_count=Count('id'))
        )

    return render(request, 'adminTripSheet.html', {
        'grouped_trips': grouped_trips
    })


def adminPrintTripSheetList(request, vehical_no, date, branch):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    # Filter trips based on VehicleNo, Date, and branch
    trips = TripSheetPrem.objects.filter(
        VehicalNo=vehical_no,
        Date=date,
        branch=branch
    )

    # Calculate total quantity
    total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

    # Aggregate data based on pay_status
    statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
    for status in statuses:
        status_trips = trips.filter(pay_status=status)
        summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
        summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
        summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
        summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
        summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
        summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

        # Update grand totals
        grand_total[status] = summary[status]['total_cost']
        grand_total['grand_freight'] += summary[status]['freight']
        grand_total['grand_hamali'] += summary[status]['hamali']
        grand_total['grand_st_charge'] += summary[status]['st_charge']
        grand_total['grand_door_charge'] += summary[status]['door_charge']
        grand_total['grand_weightAmt'] += summary[status]['weightAmt']
        grand_total['grand_total'] += summary[status]['total_cost']

    # Calculate the total value using the first row
    if trips.exists():
        first_trip = trips.first()
        total_ltr_value = float(
            first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
        total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
        total_value = total_ltr_value
    else:
        total_value = 0.0

    return render(request, 'adminPrintTripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary,
        'vehical_no': vehical_no,
        'date': date,
        'branch': branch,
    })




def exportAdminTripSheet(request, vehical_no, date, branch):
    # The branch parameter is now passed correctly via the URL
    trips = TripSheetPrem.objects.filter(
        VehicalNo=vehical_no,
        Date=date,
        branch=branch
    )

    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    # Fetch the branch name using the branch parameter
    branch_obj = get_object_or_404(Branch, companyname=branch)  # assuming 'id' is the primary key for branch

    statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
    for status in statuses:
        status_trips = trips.filter(pay_status=status)
        summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
        summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
        summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
        summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
        summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
        summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

        # Update grand totals
        grand_total[status] = summary[status]['total_cost']
        grand_total['grand_freight'] += summary[status]['freight']
        grand_total['grand_hamali'] += summary[status]['hamali']
        grand_total['grand_st_charge'] += summary[status]['st_charge']
        grand_total['grand_door_charge'] += summary[status]['door_charge']
        grand_total['grand_weightAmt'] += summary[status]['weightAmt']
        grand_total['grand_total'] += summary[status]['total_cost']

    # Create Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Admin Trip Sheet"

    # Header
    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"Branch: {branch_obj.companyname}"  # Use branch name from branch object
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = f"Driver: [Driver Name], Vehicle: {vehical_no}, Date: {date}"
    sheet["A2"].alignment = Alignment(horizontal="center")

    # Column Headers
    headers = [
        "SL/N", "LR No", "QTY", "DESC OF GOODS", "DEST", "CONSIGNOR", "CONSIGNOR GST",
        "CONSIGNEE", "CONSIGNEE GST", "TO PAY", "PAID", "CSER A/C", "CSGEE A/C"
    ]
    sheet.append(headers)

    # Trip Data
    for idx, trip in enumerate(trips, start=1):
        sheet.append([
            idx, trip.LRno, trip.qty, trip.desc, trip.dest, trip.consignor,
            trip.consignorGST, trip.consignee, trip.consigneeGST,
            trip.total_cost if trip.pay_status == "ToPay" else "",
            trip.total_cost if trip.pay_status == "Paid" else "",
            trip.total_cost if trip.pay_status == "Consigner_AC" else "",
            trip.total_cost if trip.pay_status == "Consignee_AC" else "",
        ])

    # Summary
    sheet.append([])  # Blank row
    sheet.append(["HEAD", "Weight Amt", "FREIGHT", "HAMALI", "ST CHARGE", "DOOR CHARGE", "TOTAL"])
    for status, details in summary.items():
        sheet.append([
            status, details['weightAmt'], details['freight'], details['hamali'],
            details['st_charge'], details['door_charge'], details['total_cost']
        ])

    # Grand Total
    sheet.append([])  # Blank row
    sheet.append([
        "Total", grand_total['grand_weightAmt'], grand_total['grand_freight'],
        grand_total['grand_hamali'], grand_total['grand_st_charge'],
        grand_total['grand_door_charge'], grand_total['grand_total']
    ])

    # Export File
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="trip_sheet_{vehical_no}_{date}.xlsx"'
    workbook.save(response)
    return response

def staff(request):
    if request.method == "POST":

        uid = request.session.get('username')
        branch = Branch.objects.get(email=uid)
        branchname = branch.companyname
        branchemail = branch.email

        staff = random.randint(111111, 999999)
        staffid = str(staff)

        staffname = request.POST.get('staffname')
        staffPhone = request.POST.get('staffPhone')
        staffaddress = request.POST.get('staffaddress')
        aadhar = request.POST.get('aadhar')
        passbook = request.POST.get('passbookno')

        passport = request.POST.get('passport')
        passbookphoto = request.POST.get('passport')

        passportfile = request.FILES['passport']
        fs = FileSystemStorage()
        filepassport = fs.save(passportfile.name, passportfile)
        upload_file_url = fs.url(filepassport)
        path = os.path.join(BASE_DIR, '/media/' + filepassport)

        passbookfile = request.FILES['passbook']
        fs = FileSystemStorage()
        filepassbook = fs.save(passportfile.name, passbookfile)
        upload_file_url = fs.url(filepassbook)
        path = os.path.join(BASE_DIR, '/media/' + filepassbook)

        utype = 'staff'

        if Login.objects.filter(username=staffPhone).exists():
            messages.error(request, 'Username (Phone) already exists.')
            return render(request, 'staff.html')

        Staff.objects.create(
            staffname=staffname,
            staffPhone=staffPhone,
            staffaddress=staffaddress,
            aadhar=aadhar,
            staffid=staffid,
            Branch=branchname,
            passport=passportfile,
            passbook=passbook,
            passbookphoto=passbookfile,
            branchemail=branchemail
        )
        Login.objects.create(utype=utype, username=staffPhone, password=staffid, name=staffname)

    return render(request, 'staff.html')


def view_staff(request):
    uid = request.session.get('username')
    branch = Branch.objects.get(email=uid)
    branchname = branch.companyname
    name = request.POST.get('name', '')
    if branch:
        # Filter staff data based on the branch name (case-insensitive search)
        staff_data = Staff.objects.filter(staffname__icontains=name, Branch=branchname)
    else:
        staff_data = Staff.objects.filter(Branch=branchname)
    return render(request, 'view_staff.html', {'data': staff_data})


def get_staff(request):
    query = request.GET.get('query', '')
    if query:
        staffname = Staff.objects.filter(staffname__icontains=query).values_list('staffname', flat=True)
        print('Staff Name:', list(staffname))  # Debugging: check the data in the terminal
        return JsonResponse(list(staffname), safe=False)
    return JsonResponse([], safe=False)


def delete_staff(request, pk):
    try:
        staff = Staff.objects.get(id=pk)

        user = Login.objects.filter(username=staff.staffPhone).first()
        if user:
            user.delete()
        staff.delete()

    except ObjectDoesNotExist:
        pass
    base_url = reverse('view_staff')
    return redirect(base_url)


def edit_staff(request, pk):
    # Retrieve the Staff record
    data = Staff.objects.filter(id=pk).first()  # Retrieve a single object or None

    if not data:
        return HttpResponse("Staff record not found.", status=404)

    # Store the original staffPhone
    original_staffPhone = data.staffPhone

    if request.method == "POST":
        # Get updated values from the POST request
        staffname = request.POST.get('staffname')
        staffPhone = request.POST.get('staffPhone')
        staffaddress = request.POST.get('staffaddress')
        aadhar = request.POST.get('aadhar')
        staffid = request.POST.get('staffid')

        # Update the Staff object
        data.staffname = staffname
        data.staffPhone = staffPhone
        data.staffaddress = staffaddress
        data.aadhar = aadhar
        data.staffid = staffid
        data.save()

        # Update the Login record using the original staffPhone
        user = Login.objects.filter(
            username=original_staffPhone).first()  # Fetch the user with the original phone number
        if user:
            user.username = staffPhone  # Update username to the new phone number
            user.name = staffname  # Update name
            user.password = staffid  # Update password if necessary
            user.save()

        # Redirect to a different URL after successful update
        base_url = reverse('view_staff')
        return redirect(base_url)

    return render(request, 'edit_staff.html', {'data': data})


def staffAddTripSheet(request):
    route_to = AddConsignmentTemp.objects.values_list('route_to', flat=True).distinct()
    addtrip = defaultdict(
        lambda: {'desc_product': [], 'pieces': 0, 'receiver_name': '', 'receiver_address': '', 'receiver_GST': '',
                 'sender_name': '', 'sender_GST': '', 'sender_address': '', 'pay_status': '', 'route_to': '',
                 'total': '',
                 'weightAmt': '', 'freight': '', 'hamali': '', 'door_charge': '', 'st_charge': ''}
    )

    no_data_found = False  # Flag to check if data was found

    try:
        # Fetch all consignment data without any filtering
        consignments = AddConsignmentTemp.objects.all()

        if consignments.exists():
            for consignment in consignments:
                consignment_data = addtrip[consignment.track_id]
                consignment_data['desc_product'].append(consignment.desc_product)
                consignment_data['pieces'] += consignment.pieces
                consignment_data['route_to'] = consignment.route_to
                consignment_data['receiver_name'] = consignment.receiver_name
                consignment_data['receiver_GST'] = consignment.receiver_GST
                consignment_data['receiver_address'] = consignment.receiver_address
                consignment_data['sender_name'] = consignment.sender_name
                consignment_data['sender_GST'] = consignment.sender_GST
                consignment_data['sender_address'] = consignment.sender_address
                consignment_data['pay_status'] = consignment.pay_status
                consignment_data['total_cost'] = consignment.total_cost
                consignment_data['weightAmt'] = consignment.weightAmt
                consignment_data['freight'] = consignment.freight
                consignment_data['hamali'] = consignment.hamali
                consignment_data['door_charge'] = consignment.door_charge
                consignment_data['st_charge'] = consignment.st_charge
        else:
            no_data_found = True  # Set the flag if no data is found

        # Convert the defaultdict to a list of dictionaries
        addtrip = [
            {
                'track_id': track_id,
                'desc_product': ', '.join(consignment_data['desc_product']),
                'pieces': consignment_data['pieces'],
                'route_to': consignment_data['route_to'],
                'receiver_name': consignment_data['receiver_name'],
                'receiver_GST': consignment_data['receiver_GST'],
                'receiver_address': consignment_data['receiver_address'],
                'sender_name': consignment_data['sender_name'],
                'sender_GST': consignment_data['sender_GST'],
                'sender_address': consignment_data['sender_address'],
                'pay_status': consignment_data['pay_status'],
                'total_cost': consignment_data['total_cost'],
                'weightAmt': consignment_data['weightAmt'],
                'freight': consignment_data['freight'],
                'hamali': consignment_data['hamali'],
                'door_charge': consignment_data['door_charge'],
                'st_charge': consignment_data['st_charge']
            }
            for track_id, consignment_data in addtrip.items()
        ]

    except Exception as e:
        addtrip = []
        no_data_found = True  # Set the flag if an error occurs

    return render(request, 'staffAddTripSheet.html', {
            'route_to': route_to,
            'trip': addtrip,
            'no_data_found': no_data_found  # Pass the flag to the template
        })


def staffsaveTripSheetList(request):
    print("staffsaveTripSheetList function called")
    if request.method == 'POST':
        print("POST request received")  # Debugging statement

        uid = request.session.get('username')
        if uid:
            try:
                branch = Staff.objects.get(staffPhone=uid)
                branchname = branch.Branch
                username = branch.staffname

                now = datetime.now()
                con_date = now.strftime("%Y-%m-%d")
                current_time = now.strftime("%H:%M:%S")

                total_rows = int(request.POST.get('total_rows', 0))

                selected_rows = request.POST.getlist('selected_rows')

                for i in range(1, total_rows + 1):
                    if str(i) in selected_rows:  # Only process if the row is selected
                        track_id = request.POST.get(f'track_id_{i}')
                        pieces = request.POST.get(f'pieces_{i}')
                        desc_product = request.POST.get(f'desc_product_{i}')
                        route_to = request.POST.get(f'route_to_{i}')
                        receiver_name = request.POST.get(f'receiver_name_{i}')
                        receiver_GST = request.POST.get(f'receiver_GST_{i}')
                        receiver_address = request.POST.get(f'receiver_address_{i}')
                        sender_name = request.POST.get(f'sender_name_{i}')
                        sender_GST = request.POST.get(f'sender_GST_{i}')
                        sender_address = request.POST.get(f'sender_address_{i}')
                        pay_status = request.POST.get(f'pay_status_{i}')
                        total_cost = request.POST.get(f'total_cost{i}')
                        weightAmt = request.POST.get(f'weightAmt{i}')
                        freight = request.POST.get(f'freight{i}')
                        hamali = request.POST.get(f'hamali{i}')
                        door_charge = request.POST.get(f'door_charge{i}')
                        st_charge = request.POST.get(f'st_charge{i}')

                        print(
                            f"Track ID: {track_id}, Pieces: {pieces}, Description: {desc_product}, Route: {route_to}, Receiver: {receiver_name}, Pay Status: {pay_status}, total_cost:{total_cost},weightAmt:{weightAmt},freight:{freight},hamali:{hamali},door_charge:{door_charge},st_charge:{st_charge}")  # Debugging statement

                        # Save to TripSheetTemp
                        TripSheetTemp.objects.create(
                            LRno=track_id,
                            qty=pieces,
                            desc=desc_product,
                            dest=route_to,
                            consignee=receiver_name,
                            consignor=sender_name,
                            consignorGST=sender_GST,
                            consignorAddress=sender_address,
                            consigneeGST=receiver_GST,
                            consigneeAddress=receiver_address,
                            pay_status=pay_status,
                            branch=branchname,
                            username=username,
                            Date=con_date,
                            total_cost=total_cost,
                            weightAmt=weightAmt,
                            freight=freight,
                            hamali=hamali,
                            door_charge=door_charge,
                            st_charge=st_charge,
                        )

                        # Delete from AddConsignmentTemp
                        AddConsignmentTemp.objects.filter(track_id=track_id).delete()

                        print(
                            f"Data for Track ID {track_id} saved and deleted from AddConsignmentTemp successfully.")  # Debugging statement
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
        else:
            print("No username found in session.")  # Debugging statement

        return redirect('staffAddTripSheet')  # Replace with your desired success URL

    print("Not a POST request, redirecting back to form.")  # Debugging statement
    return render(request, 'staffAddTripSheet.html')  # Redirect back to the form if not a POST request


def staffAddTripSheetList(request):
    addtrip = []  # Initialize an empty list to store trip details
    uid = request.session.get('username')
    no_data_found = False  # Flag to check if no data is found

    if uid:
        try:
            branch = Staff.objects.get(staffPhone=uid)
            user_branch = branch.Branch

            if request.method == 'POST':
                # Get the selected date from the form
                date = request.POST.get('date')

                if date:
                    # Query TripSheetTemp table based on the selected date and user's branch
                    consignments = TripSheetTemp.objects.filter(
                        Date=date,
                        branch=user_branch
                    )

                    # Check if consignments exist
                    if consignments.exists():
                        # Iterate through the results and prepare the data for the template
                        addtrip = [
                            {
                                'track_id': consignment.LRno,
                                'desc': consignment.desc,
                                'qty': consignment.qty,
                                'dest': consignment.dest,
                                'consignee': consignment.consignee,
                                'consigneeGST': consignment.consigneeGST,
                                'consigneeAddress': consignment.consigneeAddress,
                                'consignor': consignment.consignor,
                                'consignorGST': consignment.consignorGST,
                                'consignorAddress': consignment.consignorAddress,
                                'pay_status': consignment.pay_status,
                                'total_cost': consignment.total_cost,
                                'weightAmt': consignment.weightAmt,
                                'freight': consignment.freight,
                                'hamali': consignment.hamali,
                                'door_charge': consignment.door_charge,
                                'st_charge': consignment.st_charge
                            }
                            for consignment in consignments
                        ]
                    else:
                        no_data_found = True  # Set the flag if no data is found

        except Branch.DoesNotExist:
            addtrip = []
            no_data_found = True  # Set the flag if the branch does not exist

        # Render the template with the trip data and no_data_found flag
        return render(request, 'staffAddTripSheetList.html', {
            'trip': addtrip,
            'no_data_found': no_data_found,
        })


def staffSaveTripSheet(request):
    print("staffSaveTripSheet function called")
    if request.method == 'POST':
        print("POST request received")  # Debugging statement

        last_trip_id = TripSheetPrem.objects.aggregate(Max('trip_id'))['trip_id__max']
        trip_id = int(last_trip_id) + 1 if last_trip_id else 1000  # Start from a defined base if no entries exist
        con_id = str(trip_id)

        uid = request.session.get('username')
        if uid:
            try:
                branch = Staff.objects.get(staffPhone=uid)
                branchname = branch.Branch
                username = branch.staffname

                now = datetime.now()
                con_date = now.strftime("%Y-%m-%d")
                current_time = now.strftime("%H:%M:%S")

                # Get form data
                vehicle = request.POST.get('vehical')
                drivername = request.POST.get('drivername')
                adv = request.POST.get('advance')
                ltrate = request.POST.get('ltrate') or 0
                ltr = request.POST.get('liter') or 0

                literate = float(ltrate)
                liter = float(ltr)
                diesel_total = literate * liter



                # Save to Disel table
                Disel.objects.create(
                    Date=con_date,
                    vehicalno=vehicle,
                    drivername=drivername,
                    ltrate=ltrate,
                    liter=ltr,
                    total=diesel_total,  # Diesel total cost
                    trip_id=con_id
                )

                total_rows = int(request.POST.get('total_rows', 0))

                print(f"Vehicle: {vehicle}, Driver Name: {drivername}")  # Debugging statement

                for i in range(1, total_rows + 1):
                    track_id = request.POST.get(f'track_id_{i}')
                    desc = request.POST.get(f'desc_{i}')
                    qty = request.POST.get(f'qty_{i}')
                    dest = request.POST.get(f'dest_{i}')
                    consignee = request.POST.get(f'consignee_{i}')
                    consigneeGST = request.POST.get(f'consigneeGST_{i}')
                    consigneeAddress = request.POST.get(f'consigneeAddress_{i}')
                    consignor = request.POST.get(f'consignor_{i}')
                    consignorGST = request.POST.get(f'consignorGST_{i}')
                    consignorAddress = request.POST.get(f'consignorAddress_{i}')
                    total_cost = request.POST.get(f'total_cost_{i}')
                    pay_status = request.POST.get(f'pay_status_{i}')
                    weightAmt = request.POST.get(f'weightAmt_{i}')
                    freight = request.POST.get(f'freight_{i}')
                    hamali = request.POST.get(f'hamali_{i}')
                    door_charge = request.POST.get(f'door_charge_{i}')
                    st_charge = request.POST.get(f'st_charge_{i}')

                    print(
                        f"Track ID: {track_id}, Description: {desc}, Quantity: {qty}, Route: {dest}, Receiver: {consignee}")  # Debugging

                    # Save to TripSheetPrem
                    TripSheetPrem.objects.create(
                        LRno=track_id,
                        qty=qty,
                        desc=desc,
                        dest=dest,
                        consignee=consignee,
                        consigneeGST=consigneeGST,
                        consigneeAddress=consigneeAddress,
                        consignor=consignor,
                        consignorGST=consignorGST,
                        consignorAddress=consignorAddress,
                        pay_status=pay_status,
                        VehicalNo=vehicle,
                        DriverName=drivername,
                        branch=branchname,
                        username=username,
                        Date=con_date,
                        Time=current_time,
                        AdvGiven=adv,
                        LTRate=ltrate,
                        Ltr=ltr,
                        total_cost=total_cost,
                        weightAmt=float(weightAmt),
                        freight=freight,
                        hamali=hamali,
                        door_charge=door_charge,
                        st_charge=st_charge,
                        trip_id=con_id,
                        status='TripSheet Added',
                    )


                    # Delete from AddConsignmentTemp
                    TripSheetTemp.objects.filter(LRno=track_id).delete()

                    print(f"Data for Track ID {track_id} saved successfully.")  # Debugging statement


            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
            else:
                print("No username found in session.")  # Debugging statement

    return redirect('staffAddTripSheetList')  # Replace with your desired success URL

    print("Not a POST request, redirecting back to form.")  # Debugging statement
    return render(request, 'staffAddTripSheetList.html')  # Redirect back to the form if not a POST request


def staffTripSheet(request):
    return render(request, 'staffTripSheet.html')


def staffTripSheetList(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Staff.objects.get(staffPhone=uid)
            user_branch = branch.Branch

            if request.method == 'POST':
                vehicle_number = request.POST.get('vehical')
                date = request.POST.get('t3')

                if date:
                    trips = TripSheetPrem.objects.filter(
                        VehicalNo=vehicle_number,
                        Date=date,
                        branch=user_branch
                    )
                    # Calculate total quantity
                    total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

                    # Aggregate data based on pay_status
                    statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
                    for status in statuses:
                        status_trips = trips.filter(pay_status=status)
                        summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                        summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                        summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                        summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                        summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                        summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                        # Update grand totals
                        grand_total[status] = summary[status]['total_cost']
                        grand_total['grand_freight'] += summary[status]['freight']
                        grand_total['grand_hamali'] += summary[status]['hamali']
                        grand_total['grand_st_charge'] += summary[status]['st_charge']
                        grand_total['grand_door_charge'] += summary[status]['door_charge']
                        grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                        grand_total['grand_total'] += summary[status]['total_cost']

                    # Calculate the total value using the first row
                    if trips.exists():
                        first_trip = trips.first()
                        total_ltr_value = float(
                            first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                        total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                        total_value = total_ltr_value + total_adv_given
                    else:
                        total_value = 0.0

        except ObjectDoesNotExist:
            trips = TripSheetTemp.objects.none()

    return render(request, 'staffTripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary
    })


def staffViewTripSheetList(request):
    grouped_trips = []
    uid = request.session.get('username')

    if uid:
        try:
            branch = Staff.objects.get(staffPhone=uid)
            user_branch = branch.Branch

            if request.method == 'POST':
                date = request.POST.get('t3')

                if date:
                    # Group by VehicalNo and Date, and annotate with count
                    grouped_trips = (
                        TripSheetPrem.objects
                        .filter(Date=date, branch=user_branch)
                        .values('VehicalNo', 'Date')
                        .annotate(trip_count=Count('id'))
                    )

        except ObjectDoesNotExist:
            grouped_trips = []

    return render(request, 'staffViewTripSheetList.html', {
        'grouped_trips': grouped_trips
    })


def staffprintTripSheetList(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            branch = Staff.objects.get(staffPhone=uid)
            user_branch = branch.Branch

            vehical_no = request.POST.get('vehical')
            date = request.POST.get('t3')

            # Filter trips based on VehicleNo, Date, and branch
            trips = TripSheetPrem.objects.filter(
                VehicalNo=vehical_no,
                Date=date,
                branch=user_branch
            )

            # Calculate total quantity
            total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

            # Aggregate data based on pay_status
            statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
            for status in statuses:
                status_trips = trips.filter(pay_status=status)
                summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                # Update grand totals
                grand_total[status] = summary[status]['total_cost']
                grand_total['grand_freight'] += summary[status]['freight']
                grand_total['grand_hamali'] += summary[status]['hamali']
                grand_total['grand_st_charge'] += summary[status]['st_charge']
                grand_total['grand_door_charge'] += summary[status]['door_charge']
                grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                grand_total['grand_total'] += summary[status]['total_cost']

            # Calculate the total value using the first row
            if trips.exists():
                first_trip = trips.first()
                total_ltr_value = float(
                    first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                total_value = total_ltr_value
            else:
                total_value = 0.0

        except Branch.DoesNotExist:
            trips = TripSheetPrem.objects.none()  # Handle case where Branch does not exist

    return render(request, 'staffprintTripSheetList.html', {
        'trips': trips,
        'total_value': total_value,
        'total_qty': total_qty,
        'grand_total': grand_total,
        'summary': summary
    })

def staffexportTripSheetToExcel(request):
    trips = []
    total_value = 0
    total_qty = 0
    grand_total = {
        'ToPay': 0,
        'Paid': 0,
        'Consigner_AC': 0,
        'Consignee_AC': 0,
        'grand_freight': 0,
        'grand_hamali': 0,
        'grand_st_charge': 0,
        'grand_door_charge': 0,
        'grand_weightAmt': 0,
        'grand_total': 0
    }
    summary = {
        'ToPay': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Paid': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consigner_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0},
        'Consignee_AC': {'freight': 0, 'hamali': 0, 'st_charge': 0, 'door_charge': 0, 'weightAmt': 0, 'total_cost': 0}
    }

    uid = request.session.get('username')

    if uid:
        try:
            staff = Staff.objects.get(staffPhone=uid)
            user_branch = staff.Branch

            vehical_no = request.GET.get('vehical')
            date = request.GET.get('t3')

            trips = TripSheetPrem.objects.filter(
                VehicalNo=vehical_no,
                Date=date,
                branch=user_branch
            )

            total_qty = trips.aggregate(total_qty=Sum('qty'))['total_qty'] or 0

            statuses = ['ToPay', 'Paid', 'Consigner_AC', 'Consignee_AC']
            for status in statuses:
                status_trips = trips.filter(pay_status=status)
                summary[status]['freight'] = status_trips.aggregate(total=Sum('freight'))['total'] or 0
                summary[status]['hamali'] = status_trips.aggregate(total=Sum('hamali'))['total'] or 0
                summary[status]['st_charge'] = status_trips.aggregate(total=Sum('st_charge'))['total'] or 0
                summary[status]['door_charge'] = status_trips.aggregate(total=Sum('door_charge'))['total'] or 0
                summary[status]['weightAmt'] = status_trips.aggregate(total=Sum('weightAmt'))['total'] or 0
                summary[status]['total_cost'] = status_trips.aggregate(total=Sum('total_cost'))['total'] or 0

                grand_total[status] = summary[status]['total_cost']
                grand_total['grand_freight'] += summary[status]['freight']
                grand_total['grand_hamali'] += summary[status]['hamali']
                grand_total['grand_st_charge'] += summary[status]['st_charge']
                grand_total['grand_door_charge'] += summary[status]['door_charge']
                grand_total['grand_weightAmt'] += summary[status]['weightAmt']
                grand_total['grand_total'] += summary[status]['total_cost']

            if trips.exists():
                first_trip = trips.first()
                total_ltr_value = float(first_trip.LTRate * first_trip.Ltr) if first_trip.LTRate and first_trip.Ltr else 0.0
                total_adv_given = float(first_trip.AdvGiven) if first_trip.AdvGiven else 0.0
                total_value = total_ltr_value
            else:
                total_value = 0.0

        except Staff.DoesNotExist:
            trips = TripSheetPrem.objects.none()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trip Sheet"

    # Merging cells for Branch Name
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    sheet["A1"] = staff.Branch if uid else "N/A"
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Merging cells for Driver, Vehicle, and Date details
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    driver_name = first_trip.DriverName if trips.exists() else "N/A"
    sheet["A2"] = f"Driver: {driver_name}, VehicalNo: {vehical_no}, Date: {date}"
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers
    headers = [
        "SL/N", "LR No", "QTY", "DESC OF GOODS", "DEST", "CONSIGNOR", "CONSIGNOR GST",
        "CONSIGNEE", "CONSIGNEE GST", "TO PAY", "PAID", "CSER A/C", "CSGEE A/C"
    ]
    sheet.append(headers)

    # Trip Data
    for idx, trip in enumerate(trips, start=1):
        sheet.append([
            idx, trip.LRno, trip.qty, trip.desc, trip.dest, trip.consignor,
            trip.consignorGST, trip.consignee, trip.consigneeGST,
            trip.total_cost if trip.pay_status == "ToPay" else "",
            trip.total_cost if trip.pay_status == "Paid" else "",
            trip.total_cost if trip.pay_status == "Consigner_AC" else "",
            trip.total_cost if trip.pay_status == "Consignee_AC" else "",
        ])

    # Summary Table
    sheet.append([])
    sheet.append(["HEAD", "Weight Amt", "FREIGHT", "HAMALI", "ST CHARGE", "DOOR CHARGE", "TOTAL"])
    for status, details in summary.items():
        sheet.append([
            status, details['weightAmt'], details['freight'], details['hamali'],
            details['st_charge'], details['door_charge'], details['total_cost']
        ])

    # Grand Total
    sheet.append([])
    sheet.append([
        "Total", grand_total['grand_weightAmt'], grand_total['grand_freight'],
        grand_total['grand_hamali'], grand_total['grand_st_charge'],
        grand_total['grand_door_charge'], grand_total['grand_total']
    ])

    # Export File
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="trip_sheet.xlsx"'
    workbook.save(response)
    return response


def fetch_consignments(request):
    consignments = AddConsignment.objects.all()
    consignments_data = [
        {
            'id': consignment.id,
            'track_id': consignment.track_id,
            'sender_name': consignment.sender_name,
            'receiver_name': consignment.receiver_name,
        }
        for consignment in consignments
    ]
    return JsonResponse(consignments_data, safe=False)


def fetch_details(request):
    uid = request.session.get('username')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    pay_status = request.GET.get('pay_status')
    consignor_id = request.GET.get('consignor_id')
    consignee_id = request.GET.get('consignee_id')

    # Initialize an empty queryset
    consignments = AddConsignment.objects.none()
    data = []
    if uid:
        try:
            # Fetch the branch of the logged-in user
            branch = Staff.objects.get(staffPhone=uid).Branch

            # Start with filtering consignments by branch
            consignments = AddConsignment.objects.filter(branch=branch)

            # Further filter consignments based on the provided parameters
            if consignor_id:
                consignments = consignments.filter(sender_name__icontains=consignor_id)
            if consignee_id:
                consignments = consignments.filter(receiver_name__icontains=consignee_id)
            if from_date and to_date:
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                consignments = consignments.filter(date__range=(from_date, to_date))

            # Handle pay_status filtering
            if pay_status and pay_status != 'all':
                consignments = consignments.filter(pay_status__icontains=pay_status)

            # Group consignments by track_id
            grouped_data = defaultdict(lambda: {
                'track_id': '',
                'sender_name': '',
                'receiver_name': '',
                'desc_product': '',
                'pay_status': '',
                'pieces': '',
                'total_cost': 0
            })

            for consignment in consignments:
                track_id = consignment.track_id
                if track_id not in grouped_data:
                    grouped_data[track_id]['track_id'] = track_id
                    grouped_data[track_id]['sender_name'] = consignment.sender_name
                    grouped_data[track_id]['receiver_name'] = consignment.receiver_name
                    grouped_data[track_id]['pay_status'] = consignment.pay_status
                    grouped_data[track_id]['total_cost'] = consignment.total_cost

                # Concatenate pieces and desc_product as strings
                if grouped_data[track_id]['pieces']:
                    grouped_data[track_id]['pieces'] += consignment.pieces
                else:
                    grouped_data[track_id]['pieces'] = consignment.pieces

                if grouped_data[track_id]['desc_product']:
                    grouped_data[track_id]['desc_product'] += ', ' + consignment.desc_product
                else:
                    grouped_data[track_id]['desc_product'] = consignment.desc_product

            # Prepare the data for JSON response
            data = list(grouped_data.values())
        except Staff.DoesNotExist:
            print("Staff does not exist for the provided uid.")  # Handle case where Staff does not exist

    return JsonResponse({'data': data})


def branchfetch_details(request):
    uid = request.session.get('username')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    pay_status = request.GET.get('pay_status')
    consignor_id = request.GET.get('consignor_id')
    consignee_id = request.GET.get('consignee_id')

    # Initialize data and consignments
    consignments = AddConsignment.objects.none()
    data = []

    if uid:
        try:
            # Fetch the branch of the logged-in user
            branch = Branch.objects.get(email=uid)
            uname = branch.companyname

            # Start with filtering consignments by branch
            consignments = AddConsignment.objects.filter(branch=uname)

            # Further filter consignments based on the provided parameters
            if consignor_id:
                consignments = consignments.filter(sender_name__icontains=consignor_id)
            if consignee_id:
                consignments = consignments.filter(receiver_name__icontains=consignee_id)
            if from_date and to_date:
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                consignments = consignments.filter(date__range=(from_date, to_date))
            if pay_status and pay_status != 'all':
                consignments = consignments.filter(pay_status__icontains=pay_status)

            # Group consignments by track_id
            grouped_data = defaultdict(lambda: {
                'track_id': '',
                'sender_name': '',
                'receiver_name': '',
                'desc_product': '',
                'pay_status': '',
                'pieces': '',
                'total_cost': 0
            })

            for consignment in consignments:
                track_id = consignment.track_id
                if track_id not in grouped_data:
                    grouped_data[track_id]['track_id'] = track_id
                    grouped_data[track_id]['sender_name'] = consignment.sender_name
                    grouped_data[track_id]['receiver_name'] = consignment.receiver_name
                    grouped_data[track_id]['pay_status'] = consignment.pay_status
                    grouped_data[track_id]['total_cost'] = consignment.total_cost

                # Concatenate pieces and desc_product as strings
                if grouped_data[track_id]['pieces']:
                    grouped_data[track_id]['pieces'] += consignment.pieces
                else:
                    grouped_data[track_id]['pieces'] = consignment.pieces

                if grouped_data[track_id]['desc_product']:
                    grouped_data[track_id]['desc_product'] += ', ' + consignment.desc_product
                else:
                    grouped_data[track_id]['desc_product'] = consignment.desc_product

                # Sum up total costs

            # Prepare the data for JSON response
            data = list(grouped_data.values())
        except Branch.DoesNotExist:
            print("Branch does not exist for the provided uid.")  # Handle case where Branch does not exist

    # Include the uid in the response data
    return JsonResponse({'uid': uid, 'data': data})


def payment_history(request):
    return render(request, 'payment_history.html')


def credit(request):
    credit = Account.objects.all()
    return render(request, 'credit.html', {'credit': credit})


@csrf_exempt
def fetch_balance(request):
    uid = request.session.get('username')

    if uid:
        try:
            # Fetch the branch of the logged-in user
            branch = Staff.objects.get(staffPhone=uid).Branch

            if request.method == 'GET':
                sender_name = request.GET.get('sender_name')
                if sender_name:
                    # Filter accounts by sender_name and branch
                    accounts = Account.objects.filter(sender_name=sender_name, Branch=branch)
                    if accounts.exists():
                        latest_account = accounts.latest('Date')  # Get the latest record by date
                        return JsonResponse({'balance': latest_account.Balance})
                    return JsonResponse({'balance': '0'})  # Default if no records found
                return JsonResponse({'status': 'error', 'message': 'Sender name is required'})
            return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
        except Branch.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Branch does not exist for this user'})


@csrf_exempt
def submit_credit(request):
    if request.method == 'POST':
        uid = request.session.get('username')

        consignor_name = request.POST.get('consignor_name')
        credit_amount = request.POST.get('credit_amount')
        desc = request.POST.get('desc')
        now = datetime.now().replace(microsecond=0)

        if consignor_name and credit_amount:
            try:

                branch = Staff.objects.get(staffPhone=uid)
                username = branch.staffname
                branchname = branch.Branch
                # Fetch all matching records
                accounts = Account.objects.filter(sender_name=consignor_name)

                if accounts.exists():
                    # Get the latest account for calculating the new balance
                    latest_account = accounts.latest('Date')  # Assuming you want to get the latest record

                    # Calculate the new balance
                    new_balance = float(latest_account.Balance) - float(credit_amount)

                    # Create a new record with updated balance
                    new_account = Account(
                        sender_name=consignor_name,
                        credit=credit_amount,
                        debit='0',
                        TrType="ReCap",
                        particulars=desc,  # Set debit to zero
                        Balance=str(new_balance),  # Set the new balance
                        Date=now,  # Use the date of the latest record or set to current date
                        headname=username,
                        Branch=branchname
                    )
                    new_account.save()

                    return JsonResponse({'status': 'success'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'No account found with the given sender name'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        return JsonResponse({'status': 'error', 'message': 'Invalid data'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def credit_print(request):
    credit = Account.objects.all()
    return render(request, 'credit_print.html', {'credit': credit})


@csrf_exempt
def branchfetch_balance(request):
    uid = request.session.get('username')

    if uid:
        try:
            # Fetch the branch of the logged-in user
            branch = Branch.objects.get(email=uid).companyname

            if request.method == 'GET':
                sender_name = request.GET.get('sender_name')
                if sender_name:
                    # Filter accounts by sender_name and branch
                    accounts = Account.objects.filter(sender_name=sender_name, Branch=branch)
                    if accounts.exists():
                        latest_account = accounts.latest('Date')  # Get the latest record by date
                        return JsonResponse({'balance': latest_account.Balance})
                    return JsonResponse({'balance': '0'})  # Default if no records found
                return JsonResponse({'status': 'error', 'message': 'Sender name is required'})
            return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
        except Branch.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Branch does not exist for this user'})


def branchPaymenyHistory(request):
    return render(request, 'branchPaymenyHistory.html')


def branchcredit(request):
    credit = Account.objects.all()
    return render(request, 'branchcredit.html', {'credit': credit})


import logging

logger = logging.getLogger(__name__)


@csrf_exempt
def branchsubmit_credit(request):
    if request.method == 'POST':
        uid = request.session.get('username')

        consignor_name = request.POST.get('consignor_name')
        credit_amount = request.POST.get('credit_amount')
        desc = request.POST.get('desc')
        now = datetime.now().replace(microsecond=0)

        if consignor_name and credit_amount:
            try:

                branch = Branch.objects.get(email=uid)
                username = branch.headname
                branchcompany = branch.companyname
                # Fetch all matching records
                accounts = Account.objects.filter(sender_name=consignor_name)

                if accounts.exists():
                    # Get the latest account for calculating the new balance
                    latest_account = accounts.latest('Date')  # Assuming you want to get the latest record

                    # Calculate the new balance
                    new_balance = float(latest_account.Balance) - float(credit_amount)

                    # Create a new record with updated balance
                    new_account = Account(
                        sender_name=consignor_name,
                        credit=credit_amount,
                        debit='0',
                        TrType="ReCap",
                        particulars=desc,  # Set debit to zero
                        Balance=str(new_balance),  # Set the new balance
                        Date=now,  # Use the date of the latest record or set to current date
                        headname=username,
                        Branch=branchcompany
                    )
                    new_account.save()

                    return JsonResponse({'status': 'success'})
                else:
                    return JsonResponse({'status': 'error', 'message': 'No account found with the given sender name'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        return JsonResponse({'status': 'error', 'message': 'Invalid data'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def branchcredit_print(request):
    credit = Account.objects.all()
    return render(request, 'branchcredit_print.html', {'credit': credit})


def staffcredit_print(request):
    credit = Account.objects.all()
    return render(request, 'staffcredit_print.html', {'credit': credit})


# Set up logging
import logging

logger = logging.getLogger(__name__)


def branchfetch_account_details(request):
    if request.method == 'POST':
        uid = request.session.get('username')

        sender_name = request.POST.get('sender_name')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        logger.info(f"Received request with sender_name: {sender_name}, from_date: {from_date}, to_date: {to_date}")

        # Check if the required parameters are provided
        if sender_name and from_date and to_date:
            try:
                branch = Branch.objects.get(email=uid).companyname

                # Convert from_date and to_date to proper datetime objects
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

                # Ensure the end date includes the entire day
                to_date_end = to_date + timedelta(days=1)

                # Fetch all accounts based on sender_name, branch, and date range
                accounts = Account.objects.filter(
                    sender_name=sender_name,
                    Branch=branch,
                    Date__gte=from_date,
                    Date__lt=to_date_end
                ).values(
                    'Date', 'track_number', 'TrType', 'particulars', 'debit', 'credit', 'Balance'
                ).order_by('Date')  # Order by date if needed

                logger.info(f"Fetched accounts: {list(accounts)}")

                return render(request, 'branchcredit_print.html', {
                    'accounts': accounts,
                    'sender_name': sender_name,
                    'from_date_str': from_date,
                    'to_date_str': to_date,
                    'branch': branch
                })

            except ValueError:
                logger.error("Invalid date format")
                return render(request, 'branchcredit_print.html', {'error': 'Invalid date format'})

    logger.error("Missing required parameters")
    return render(request, 'branchcredit_print.html', {'error': 'Missing required parameters'})


logger = logging.getLogger(__name__)


@csrf_exempt
def fetch_account_details(request):
    if request.method == 'POST':

        uid = request.session.get('username')

        sender_name = request.POST.get('sender_name')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        logger.info(f"Received request with sender_name: {sender_name}, from_date: {from_date}, to_date: {to_date}")

        # Check if the required parameters are provided
        if sender_name and from_date and to_date:
            try:
                branch = Staff.objects.get(staffPhone=uid).Branch

                # Convert from_date and to_date to proper datetime objects
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

                # Ensure the end date includes the entire day
                to_date_end = to_date + timedelta(days=1)

                # Fetch all accounts based on sender_name, branch, and date range
                accounts = Account.objects.filter(
                    sender_name=sender_name,
                    Branch=branch,
                    Date__gte=from_date,
                    Date__lt=to_date_end
                ).values(
                    'Date', 'track_number', 'TrType', 'particulars', 'debit', 'credit', 'Balance'
                ).order_by('Date')  # Order by date if needed

                logger.info(f"Fetched accounts: {list(accounts)}")

                return render(request, 'staffcredit_print.html', {
                    'accounts': accounts,
                    'sender_name': sender_name,
                    'from_date_str': from_date,
                    'to_date_str': to_date,
                    'branch': branch
                })

            except ValueError:
                logger.error("Invalid date format")
                return render(request, 'staffcredit_print.html', {'error': 'Invalid date format'})

    logger.error("Missing required parameters")
    return render(request, 'staffcredit_print.html', {'error': 'Missing required parameters'})


def branchExpenses(request):
    return render(request, 'branchExpenses.html')


def savebranchExpenses(request):
    if request.method == 'POST':
        uid = request.session.get('username')
        if uid:
            try:
                branch = Branch.objects.get(email=uid)
                branchname = branch.companyname
                username = branch.headname

                # Parse and validate date
                date_str = request.POST.get('date')
                try:
                    date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    print("Invalid date format.")  # Debugging statement
                    return redirect('branchExpenses')

                # Parse and validate amount
                amount = request.POST.get('amt')
                reason = request.POST.get('reason')

                Expenses.objects.create(
                    Date=date,
                    Reason=reason,
                    Amount=amount,

                    username=username,
                    branch=branchname
                )
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
        else:
            print("No username found in session.")  # Debugging statement

        return redirect('branchExpenses')  # Replace with your desired success URL

    return render(request, 'branchExpenses.html')


def branchViewExpenses(request):
    expenses = []
    if request.method == 'POST':
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        uid = request.session.get('username')
        if uid:
            try:
                branch = Branch.objects.get(email=uid)  # Get the branch for the logged-in user
                branch_name = branch.companyname  # Assuming companyname is used as the branch identifier

                if from_date_str and to_date_str:
                    try:
                        # Parse the date strings into datetime objects
                        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

                        # Fetch expenses within the specified date range and for the logged-in branch
                        expenses = Expenses.objects.filter(
                            Date__range=(from_date, to_date),
                            branch=branch_name
                        )

                    except ValueError:
                        print("Invalid date format.")  # Handle invalid date formats
                else:
                    print("Both from_date and to_date are required.")
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Handle the case where the branch is not found

    return render(request, 'branchViewExpenses.html', {'expenses': expenses})


def adminExpenses(request):
    return render(request, 'adminExpenses.html')


def saveadminExpenses(request):
    if request.method == 'POST':
        uid = request.session.get('username')
        if uid:
            try:
                branch = Login.objects.get(username=uid)
                branchname = branch.utype
                username = branch.name

                # Parse and validate date
                date_str = request.POST.get('date')
                try:
                    date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    print("Invalid date format.")  # Debugging statement
                    return redirect('adminExpenses')

                amount = request.POST.get('amt')
                reason = request.POST.get('reason')
                salaryDetails = request.POST.get('salaryDetails')

                Expenses.objects.create(
                    Date=date,
                    Reason=reason,
                    Amount=amount,
                    staffname=salaryDetails,
                    username=username,
                    branch=branchname
                )
            except Branch.DoesNotExist:
                print("Branch does not exist.")  # Debugging statement
        else:
            print("No username found in session.")  # Debugging statement

        return redirect('adminExpenses')  # Replace with your desired success URL

    return render(request, 'adminExpenses.html')


def adminViewExpenses(request):
    expenses = []
    if request.method == 'POST':
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        if from_date_str and to_date_str:
            try:
                # Parse the date strings into datetime objects
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

                expenses = Expenses.objects.filter(Date__range=(from_date, to_date))

            except ValueError:
                print("Invalid date format.")  # Handle invalid date formats
        else:
            print("Both from_date and to_date are required.")
    return render(request, 'adminViewExpenses.html', {'expenses': expenses})


def branchConsignorView(request):
    uid = request.session.get('username')
    if uid:
        branch = Branch.objects.get(email=uid)
        branchname = branch.companyname
        consignor = Consignor.objects.filter(branch=branchname)
    return render(request, 'branchConsignorView.html', {'consignor': consignor})


def branchConsigneeView(request):
    uid = request.session.get('username')
    if uid:
        branch = Branch.objects.get(email=uid)
        branchname = branch.companyname
        consignee = Consignee.objects.filter(branch=branchname)
    return render(request, 'branchConsigneeView.html', {'consignee': consignee})


def adminConsignorView(request):
    consignor = []  # Initialize consignee as an empty list

    if request.method == 'POST':
        branch = request.POST.get('t2')
        print(f"Branch: {branch}")  # Debugging: Print the branch name
        consignor = Consignor.objects.filter(branch=branch)
        print(f"Consignee: {consignor}")  # Debugging: Print the consignee queryset

    return render(request, 'adminConsignorView.html', {'consignor': consignor})


def adminConsigneeView(request):
    consignee = []  # Initialize consignee as an empty list

    if request.method == 'POST':
        branch = request.POST.get('t2')
        print(f"Branch: {branch}")  # Debugging: Print the branch name
        consignee = Consignee.objects.filter(branch=branch)
        print(f"Consignee: {consignee}")  # Debugging: Print the consignee queryset

    return render(request, 'adminConsigneeView.html', {'consignee': consignee})


def adminstaff_view(request):
    branch = request.POST.get('branch', '')
    if branch:
        # Filter staff data based on the branch name (case-insensitive search)
        staff_data = Staff.objects.filter(Branch__icontains=branch)
    else:
        # If no branch is provided, fetch all staff data
        staff_data = Staff.objects.all()

    # Render the template with the filtered data
    return render(request, 'adminstaff_view.html', {'data': staff_data, 'branch': branch})


from django.utils.dateparse import parse_date


def adminView_Consignment(request):
    grouped_userdata = {}  # Initialize as an empty dictionary to group data

    # Fetch all consignments initially
    queryset = AddConsignment.objects.all()

    if request.method == 'POST':
        # Get filter criteria from the POST data
        branch = request.POST.get('t2')
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')
        consigner = request.POST.get('consigner')
        consignee = request.POST.get('consignee')
        track_id = request.POST.get('lrno')

        # Parse dates
        from_date = parse_date(from_date_str) if from_date_str else None
        to_date = parse_date(to_date_str) if to_date_str else None

        # Apply filters only if there is a POST request
        if branch:
            queryset = queryset.filter(branch=branch)
        if consigner:
            queryset = queryset.filter(sender_name=consigner)
        if consignee:
            queryset = queryset.filter(receiver_name=consignee)
        if track_id:
            queryset = queryset.filter(track_id=track_id)

        if from_date and to_date:
            queryset = queryset.filter(date__range=(from_date, to_date))
        elif from_date:
            queryset = queryset.filter(date__gte=from_date)
        elif to_date:
            queryset = queryset.filter(date__lte=to_date)

    # Group consignments by track_id and concatenate product details
    for consignment in queryset:
        track_id = consignment.track_id
        if track_id not in grouped_userdata:
            grouped_userdata[track_id] = {
                'branch': consignment.branch,
                'route_from': consignment.route_from,
                'route_to': consignment.route_to,
                'sender_name': consignment.sender_name,
                'sender_mobile': consignment.sender_mobile,
                'receiver_name': consignment.receiver_name,
                'receiver_mobile': consignment.receiver_mobile,
                'total_cost': 0,
                'pieces': 0,
                'weight': consignment.weight,
                'pay_status': consignment.pay_status,
                'consignment_status': consignment.consignment_status,
                'products': []
            }
        # Aggregate total cost and pieces
        grouped_userdata[track_id]['total_cost'] += consignment.total_cost
        grouped_userdata[track_id]['pieces'] += consignment.pieces

        # Concatenate product details without ID
        product_detail = consignment.desc_product
        grouped_userdata[track_id]['products'].append(product_detail)

    # Convert the list of product details to a single string
    for track_id, details in grouped_userdata.items():
        details['products'] = ', '.join(details['products'])

    return render(request, 'adminView_Consignment.html', {'grouped_userdata': grouped_userdata})


def toggle_consignment_status(request, track_id):
    # Fetch the consignment by track_id
    consignment = get_object_or_404(AddConsignment, track_id=track_id)

    # Toggle the status between 'Completed' and 'Pending'
    if consignment.consignment_status == 'Complete':
        consignment.consignment_status = 'Pending'
    else:
        consignment.consignment_status = 'Complete'

    # Save the updated status
    consignment.save()

    # Redirect back to the consignment status page
    return redirect('adminView_Consignment')


def admininvoiceConsignment(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity

    try:
        # Filter consignments by track_id
        consignments = AddConsignment.objects.filter(track_id=track_id)
        # Get common details from the first consignment
        consignment = consignments.first()

        # Fetch the branch name from the consignment
        branch_name = consignment.branch  # Adjust this field based on your model
        branchemail = consignment.branchemail
        # Fetch branch details using the branch name
        branchdetails = get_object_or_404(Branch, email=branchemail)

        if not consignments.exists():
            return render(request, '404.html')  # Handle case where no consignments are found.
        branchdetails.services = mark_safe(branchdetails.services.replace(',', '<br/>'))


        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'desc_product': consignment.desc_product,

            }
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)

            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    return render(request, 'admininvoiceConsignment.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),  # Include the aggregated copy types
        'totalqty': totalqty  # Pass total quantity to the template

    })


def staffinvoiceConsignment(request, track_id):
    # Filter consignments by track_id
    consignments = AddConsignment.objects.filter(track_id=track_id)
    uid = request.session.get('username')
    branchname = Staff.objects.get(staffPhone=uid)
    branch = branchname.Branch
    branchdetails = Branch.objects.get(companyname=branch)

    if not consignments.exists():
        return render(request, '404.html')  # Handle the case where no consignments are found.

    # Get common details from the first consignment
    consignment = consignments.first()

    # Collect copy_types from all consignments with the same track_id
    copy_types = consignments.values_list('copy_type', flat=True).distinct()

    # Convert the queryset to a list and join them into a single string for display
    copy_type_list = ', '.join(copy_types)

    # Pass the first consignment, the entire list of items, and the copy types to the template
    return render(request, 'staffinvoiceConsignment.html', {
        'consignment': consignment,
        'items': consignments,
        'branchdetails': branchdetails,
        'copy_types': copy_type_list  # Include the aggregated copy types
    })


def partywise_list(request):
    # Get filter values from the request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    sender_name = request.GET.get('sender_name')
    receiver_name = request.GET.get('consignee')

    # Start building the query
    queryset = AddConsignment.objects.all()

    # Apply date range filter if both from_date and to_date are provided
    if from_date_str and to_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

            # Filter the queryset by the date range
            queryset = queryset.filter(date__range=(from_date, to_date))
        except ValueError:
            return render(request, 'partywise_report.html', {
                'error': 'Invalid date format.'
            })

    # Apply sender_name filter if provided
    if sender_name:
        queryset = queryset.filter(sender_name__icontains=sender_name)
    if receiver_name:
        queryset = queryset.filter(receiver_name__icontains=receiver_name)

    # Group by sender_name and calculate sum of pieces, total cost, and count of track_id
    consignments_by_sender = queryset.values('sender_name').annotate(
        total_pieces=Sum('pieces'),
        total_cost=Sum('total_cost'),
        track_id_count=Count('track_id', distinct=True)
    ).order_by('sender_name')

    # Pass the aggregated data to the template
    context = {
        'consignments_by_sender': consignments_by_sender,
        'from_date': from_date_str,
        'to_date': to_date_str,
        'sender_name': sender_name, }

    return render(request, 'partywise_report.html', context)


def partywise_detail(request, sender_name):
    # Get filter values from the request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    # Start building the query
    consignments = AddConsignment.objects.filter(sender_name=sender_name)

    # Apply date range filter if both from_date and to_date are provided
    if from_date_str and to_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

            # Filter the queryset by the date range
            consignments = consignments.filter(date__range=(from_date, to_date))
        except ValueError:
            return render(request, 'partywise_detail.html', {
                'error': 'Invalid date format.',
                'sender_name': sender_name,
            })

    if not consignments.exists():
        return render(request, 'partywise_detail.html', {'error': 'No consignments found for this sender.'})

    # Aggregate details based on Consignment_id
    aggregated_data = consignments.values(
        'Consignment_id',
        'track_id',
        'sender_name',
        'sender_mobile',
        'sender_address',
        'receiver_name',
        'receiver_mobile',
        'receiver_address',
        'date',
        'route_from',
        'route_to',
        'prod_invoice',
        'prod_price',
        'branch',
        'name',
        'time',
        'copy_type',
        'delivery',
        'eway_bill'
    ).annotate(
        total_cost=Sum('total_cost'),
        pieces=Sum('pieces'),
        weight=Sum('weight'),
        freight=Sum('freight'),
        hamali=Sum('hamali'),
        door_charge=Sum('door_charge'),
        st_charge=Sum('st_charge'),
        weightAmt=Sum('weightAmt'),
    ).order_by('Consignment_id')

    # Create a list of dictionaries for the final data to be displayed
    detailed_data = []
    for consignment in aggregated_data:
        descriptions = consignments.filter(Consignment_id=consignment['Consignment_id']).values_list('desc_product',
                                                                                                     flat=True)
        # Append each description with aggregated data
        detailed_data.append({
            **consignment,
            'desc_products': descriptions
        })

    # Calculate total pieces for the sender
    total_pieces = consignments.aggregate(total_pieces=Sum('pieces'))['total_pieces'] or 0

    return render(request, 'partywise_detail.html', {
        'sender_name': sender_name,
        'consignments': detailed_data,
        'total_pieces': total_pieces
    })


def disel_report(request):
    # Retrieve date parameters from the GET request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    # Convert date strings to datetime objects
    if from_date_str and to_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            # Handle incorrect date format
            return render(request, 'disel_report.html', {
                'data': [],
                'total_litres': 0,
                'total_amount': 0,
                'error_message': 'Invalid date format. Use YYYY-MM-DD.'
            })

        # Filter data based on date range and total > 0
        data = Disel.objects.filter(Date__range=[from_date, to_date], total__gt=0)
    else:
        # If no dates are provided, show all data where total > 0
        data = Disel.objects.filter(total__gt=0)

    # Calculate the total litres and amount
    total_litres = data.aggregate(Sum('liter'))['liter__sum'] or 0
    total_amount = data.aggregate(Sum('total'))['total__sum'] or 0

    # Pass data and totals to the template
    return render(request, 'disel_report.html', {
        'data': data,
        'total_litres': total_litres,
        'total_amount': total_amount,
        'error_message': ''
    })


def account_report(request):
    credit = Account.objects.all()
    return render(request, 'account_report.html', {'credit': credit})


@csrf_exempt
def adminfetch_account_details(request):
    if request.method == 'POST':

        sender_name = request.POST.get('sender_name')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        logger.info(f"Received request with sender_name: {sender_name}, from_date: {from_date}, to_date: {to_date}")

        # Check if the required parameters are provided
        if sender_name and from_date and to_date:
            try:

                # Convert from_date and to_date to proper datetime objects
                from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

                # Ensure the end date includes the entire day
                to_date_end = to_date + timedelta(days=1)

                # Fetch all accounts based on sender_name, branch, and date range
                accounts = Account.objects.filter(
                    sender_name=sender_name,
                    Date__gte=from_date,
                    Date__lt=to_date_end
                ).values(
                    'Date', 'track_number', 'TrType', 'particulars', 'debit', 'credit', 'Balance'
                ).order_by('Date')  # Order by date if needed

                logger.info(f"Fetched accounts: {list(accounts)}")

                return render(request, 'account_report.html', {
                    'accounts': accounts,
                    'sender_name': sender_name,
                    'from_date_str': from_date,
                    'to_date_str': to_date,
                    'branch': branch
                })

            except ValueError:
                logger.error("Invalid date format")
                return render(request, 'account_report.html', {'error': 'Invalid date format'})

    logger.error("Missing required parameters")
    return render(request, 'account_report.html', {'error': 'Missing required parameters'})


def get_account_details(request):
    branch = request.GET.get('branch', '')
    if branch:
        accounts = Account.objects.filter(Branch__icontains=branch)
        accounts_data = list(
            accounts.values('track_number', 'sender_name', 'Branch', 'headname', 'TrType', 'debit', 'credit',
                            'Balance'))
        return JsonResponse(accounts_data, safe=False)
    return JsonResponse([], safe=False)


def unloaded_LR_report(request):
    # Extract query parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    route_to = request.GET.get('dest')  # Get route_to/destination filter

    # Initialize variables for start_date and end_date
    start_date = None
    end_date = None

    # Convert string dates to datetime objects
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            pass  # Handle invalid date format if necessary

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            pass  # Handle invalid date format if necessary

    # If end_date is provided, extend it to the end of the day
    if end_date:
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Filter consignments based on date range and route_to
    consignments = AddConsignmentTemp.objects.all()

    if start_date:
        consignments = consignments.filter(date__gte=start_date)

    if end_date:
        consignments = consignments.filter(date__lte=end_date)

    if route_to:  # Apply filter for route_to if provided
        consignments = consignments.filter(route_to__icontains=route_to)

    # Render the template with the filtered consignments
    return render(request, 'unloaded_LR_report.html', {'consignments': consignments})


def advance_report(request):
    driver_name = request.GET.get('driver_name')
    vehicalno = request.GET.get('vehicalno')  # Fixed name to match form
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    # Convert date strings to date objects
    from_date = datetime.strptime(from_date_str, '%Y-%m-%d') if from_date_str else None
    to_date = datetime.strptime(to_date_str, '%Y-%m-%d') if to_date_str else None

    # Initialize the filters dictionary
    filters = {}
    if driver_name:
        filters['DriverName__iexact'] = driver_name  # Case-insensitive driver name filter
    if vehicalno:
        filters['VehicalNo__iexact'] = vehicalno  # Case-insensitive vehicle number filter
    if from_date and to_date:
        filters['Date__range'] = [from_date, to_date]

    # Add the condition for AdvGiven to be more than 0
    filters['AdvGiven__gt'] = 0

    # Fetch the results based on the filters and group by trip_id
    results = TripSheetPrem.objects.filter(**filters).values(
        'trip_id', 'VehicalNo', 'DriverName', 'AdvGiven', 'Date'  # Include the fields you need
    ).annotate(
        total_advances=Count('AdvGiven')
    ).order_by('trip_id')

    if not results:
        print("No results found")

    return render(request, 'advance_report.html', {
        'results': results,
        'vehicalno': vehicalno,
        'driver_name': driver_name,
        'from_date': from_date_str,
        'to_date': to_date_str
    })


def profit_report(request):
    # Get the from_date and to_date from the request (if provided)
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None

    # Query all consignments and expenses
    consignments = AddConsignment.objects.all()
    expenses = Expenses.objects.all()

    # Filter by date range if provided
    if from_date and to_date:
        consignments = consignments.filter(date__range=[from_date, to_date])
        expenses = expenses.filter(Date__range=[from_date, to_date])

    # Track already processed track_ids
    processed_track_ids = set()

    # This list will store the unique consignments
    unique_consignments = []

    # Iterate over all consignments to ensure only unique track_id costs are added
    for consignment in consignments:
        track_id = consignment.track_id
        # If track_id is not already processed, add its total_cost to the unique list
        if track_id not in processed_track_ids:
            processed_track_ids.add(track_id)
            unique_consignments.append(consignment)

    # Now group by date and branch, summing the total_cost of the unique consignments
    consignments_grouped = (
        AddConsignment.objects.filter(id__in=[c.id for c in unique_consignments])
        .values('date', 'branch')
        .annotate(total_cost=Sum('total_cost'))
        .order_by('date', 'branch')
    )

    # Group expenses by date and branch, and calculate total Amount for each group
    expenses_grouped = expenses.values('Date', 'branch').annotate(
        total_amount=Sum('Amount')
    ).order_by('Date', 'branch')

    # Calculate grand totals for consignments and expenses
    grand_total_consignment = sum(item['total_cost'] for item in consignments_grouped)
    grand_total_expenses = sum(item['total_amount'] for item in expenses_grouped)

    # Calculate combined grand total
    combined_grand_total = grand_total_consignment + grand_total_expenses

    # Calculate profit or loss
    total_balance = grand_total_consignment - grand_total_expenses

    # Set profit and loss
    profit = total_balance if total_balance > 0 else 0
    loss = abs(total_balance) if total_balance < 0 else 0

    # Pass the grouped data and totals to the template
    return render(request, 'profit_report.html', {
        'consignments': consignments_grouped,
        'expenses': expenses_grouped,
        'grand_total_consignment': grand_total_consignment,
        'grand_total_expenses': grand_total_expenses,
        'combined_grand_total': combined_grand_total,
        'profit': profit,
        'loss': loss,
        'from_date': from_date_str,
        'to_date': to_date_str,
    })


def get_balance(request):
    cust_id = request.GET.get('cust_id')  # Get cust_id from the query parameters
    if cust_id:
        try:
            # Fetch the most recent balance for the given cust_id
            collections = Collection.objects.filter(cust_id=cust_id, consignment_status='Complete')

            # Sum all balances for the given cust_id
            total_balance = collections.aggregate(Sum('balance'))['balance__sum'] or 0

            return JsonResponse({'balance': str(total_balance)})
        except Exception as e:  # Catch any unexpected exceptions
            return JsonResponse({'balance': None, 'error': str(e)})
    return JsonResponse({'balance': None})  # Return None if cust_id is not provided


def consignmentStatus(request):
    grouped_userdata = {}  # Initialize as an empty dictionary to group data

    # Fetch all 'Pending' consignments by default
    queryset = AddConsignment.objects.filter(consignment_status='Pending')

    # Check if it's a POST request to apply date filtering
    if request.method == 'POST':
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        # Parse dates
        from_date = parse_date(from_date_str) if from_date_str else None
        to_date = parse_date(to_date_str) if to_date_str else None

        # Apply date range filtering if dates are provided
        if from_date and to_date:
            queryset = queryset.filter(date__range=(from_date, to_date))
        elif from_date:
            queryset = queryset.filter(date__gte=from_date)
        elif to_date:
            queryset = queryset.filter(date__lte=to_date)

    # Group consignments by track_id and concatenate product details
    for consignment in queryset:
        track_id = consignment.track_id
        if track_id not in grouped_userdata:
            grouped_userdata[track_id] = {
                'branch': consignment.branch,
                'route_from': consignment.route_from,
                'route_to': consignment.route_to,
                'sender_name': consignment.sender_name,
                'sender_mobile': consignment.sender_mobile,
                'receiver_name': consignment.receiver_name,
                'receiver_mobile': consignment.receiver_mobile,
                'total_cost': 0,
                'pieces': 0,
                'weight': consignment.weight,
                'pay_status': consignment.pay_status,
                'consignment_status': consignment.consignment_status,
                'products': []
            }
        # Aggregate total cost and pieces
        grouped_userdata[track_id]['total_cost'] += consignment.total_cost
        grouped_userdata[track_id]['pieces'] += consignment.pieces

        # Concatenate product details
        product_detail = consignment.desc_product
        grouped_userdata[track_id]['products'].append(product_detail)

    # Convert the list of product details to a single string
    for track_id, details in grouped_userdata.items():
        details['products'] = ', '.join(details['products'])

    return render(request, 'consignmentStatus.html', {'grouped_userdata': grouped_userdata})


def complete_consignment(request, track_id):
    # Get the consignment object using the track_id
    consignment = get_object_or_404(AddConsignment, track_id=track_id)

    # Update the consignment_status to "Complete"
    consignment.consignment_status = "Complete"
    consignment.save()  # Save the changes to the AddConsignment model

    # Check if the track_id exists in the Collection table
    try:
        collection = Collection.objects.get(lrNo=track_id)
        # If found, update the consignment_status in the Collection table
        collection.consignment_status = "Complete"
        collection.save()  # Save the changes to the Collection model
    except Collection.DoesNotExist:
        # If the Collection entry does not exist, do nothing
        pass

    # Redirect or respond back after completion
    return redirect('consignmentStatus')  # Replace 'consignmentStatus' with the correct URL name


def collection(request):
    grouped_userdata = {}  # Initialize as an empty dictionary to group data

    # Default data retrieval
    con = AddConsignment.objects.all()
    track_ids = con.values_list('track_id', flat=True)  # Fetch track_id from consignment queryset

    # Initialize the queryset to get Collection objects where balance > 0 and track_id matches
    queryset = Collection.objects.filter(balance__gt=0, lrNo__in=track_ids)

    # Check if the request method is POST (i.e., when the user submits the date filter)
    if request.method == 'POST':
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        # Parse dates if provided
        from_date = parse_date(from_date_str) if from_date_str else None
        to_date = parse_date(to_date_str) if to_date_str else None

        print(f"From Date: {from_date}")  # Debugging: Print the from date
        print(f"To Date: {to_date}")  # Debugging: Print the to date

        # Apply the date filter to the queryset
        if from_date and to_date:
            queryset = queryset.filter(date__range=(from_date, to_date))
        elif from_date:
            queryset = queryset.filter(date__gte=from_date)
        elif to_date:
            queryset = queryset.filter(date__lte=to_date)

        print(f"Filtered Consignments: {queryset}")  # Debugging: Print the filtered queryset

    # Group consignments by track_id and aggregate required fields
    for consignment in queryset:
        lrNo = consignment.lrNo
        if lrNo not in grouped_userdata:
            grouped_userdata[lrNo] = {
                'branch': consignment.branch,
                'lrNo': consignment.lrNo,
                'sender_name': consignment.sender_name,
                'pay_status': consignment.pay_status,
                'total': consignment.total,
                'amount': consignment.amount,
                'balance': consignment.balance,
                'consignment_status': consignment.consignment_status,
            }

    # Render the template and pass the grouped data
    return render(request, 'collection.html', {'grouped_userdata': grouped_userdata})


def save_payment(request):
    if request.method == 'POST':
        try:
            # Parse the incoming JSON request body
            data = json.loads(request.body)
            track_id = data.get('track_id')
            amount = data.get('amount')
            desc = data.get('desc')
            prev_amount =data.get('amount')

            # Validate amount
            if amount is None or amount == '':
                return JsonResponse({'success': False, 'error': 'Amount is required'})


            amount = float(amount)  # Convert to float after validation

            # Fetch the existing collection entry by track_id
            collection = Collection.objects.get(lrNo=track_id)
            sendername = collection.sender_name

            # Update the amount and balance for Collection
            collection.amount += amount  # Assuming amount is cumulative
            collection.balance -= amount  # Reduce balance by the paid amount
            collection.desc = desc  # Reduce balance by the paid amount
            collection.save()

            return JsonResponse({'success': True})

        except Collection.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Collection entry not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request'})

import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .models import Collection
def collection_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Filter data by `desc` and optionally by date range
    collections = Collection.objects.filter(desc__isnull=False).exclude(desc="")
    if from_date and to_date:
        collections = collections.filter(date__range=[from_date, to_date])
    elif from_date:
        collections = collections.filter(date__gte=from_date)
    elif to_date:
        collections = collections.filter(date__lte=to_date)

    # Calculate grand totals for total and amount
    grand_total = sum(collection.total for collection in collections)
    grand_amount = sum(collection.amount for collection in collections)

    if request.GET.get('export') == 'excel':
        # Create Excel report with filtered data
        return export_collection_to_excel(collections, grand_total, grand_amount)

    return render(request, 'collectionReport.html', {
        'collections': collections,
        'grand_total': grand_total,
        'grand_amount': grand_amount,
        'from_date': from_date,
        'to_date': to_date,
    })


def export_collection_to_excel(collections, grand_total, grand_amount):
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection Report"

    # Add headers (excluding 'route_from', 'route_to', 'pay_status')
    headers = ["LR No", "Sender Name", "Receiver Name", "Total", "Date", "Description", "Amount"]
    ws.append(headers)

    # Add data rows
    for collection in collections:
        ws.append([
            collection.lrNo, collection.sender_name, collection.receiver_name,
            collection.total, collection.date, collection.desc, collection.amount
        ])
    ws.append([])

    # Add grand total row
    ws.append(["", "", "", grand_total, "", "", grand_amount])

    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Collection_Report.xlsx'
    wb.save(response)
    return response



def cancel_trip(request, trip_id):
    trip = get_object_or_404(TripSheetPrem, LRno=trip_id)
    con = get_object_or_404(AddConsignment, track_id=trip_id)

    # Copy the trip data to TripSheetTemp
    AddConsignmentTemp.objects.create(
        track_id=con.track_id,
        Consignment_id=con.Consignment_id,
        sender_name=con.sender_name,
        sender_mobile=con.sender_mobile,
        sender_address=con.sender_address,
        sender_GST=con.sender_GST,
        receiver_name=con.receiver_name,
        receiver_mobile=con.receiver_mobile,
        receiver_address=con.receiver_address,
        receiver_GST=con.receiver_GST,
        desc_product=con.desc_product,
        pieces=con.pieces,
        prod_invoice=con.prod_invoice,
        prod_price=con.prod_price,
        weightAmt=con.weightAmt,
        weight=con.weight,
        balance=con.balance,
        freight=con.freight,
        hamali=con.hamali,
        door_charge=con.door_charge,
        st_charge=con.st_charge,
        route_from=con.route_from,
        route_to=con.route_to,
        total_cost=con.total_cost,
        date=con.date,
        pay_status=con.pay_status,
        branch=con.branch,
        name=con.name,
        time=con.time,
        copy_type=con.copy_type,
        delivery=con.delivery,
        eway_bill=con.eway_bill
    )

    # Optionally, mark the trip as cancelled (if needed for other logic)
    trip.is_cancelled = True
    trip.save()

    # Remove the trip record from TripSheetPrem after saving
    trip.delete()

    messages.success(request,
                     f"Trip {trip.LRno} has been cancelled, saved in TripSheetTemp, and removed from TripSheetPrem.")
    return redirect('viewTripSheetList')  # Replace with the actual view name if needed


def customerLogin(request):
    if request.method == 'POST':
        # Get the email from the frontend
        email = request.POST.get('email')

        # Check if the email exists in the UserLogin table
        try:
            user = UserLogin.objects.get(email=email)
            # Save email in session
            request.session['user_email'] = email
            # Redirect to CustomerHome page
            return redirect('customerHome')  # Replace 'customer_home' with your URL name for the CustomerHome page
        except UserLogin.DoesNotExist:
            # Email does not exist, show an error message
            messages.error(request, "Email not found. Please try again.")
            return render(request, 'customerLogin.html')  # Render the login page again

    return render(request, 'customerLogin.html')


def customerHome(request):
    return render(request, 'customerHome.html')


def customerConsignment(request):
    uid = request.session.get('user_email')  # Get email from session
    grouped_userdata = {}

    if uid:
        try:
            # First, check if the email is in the Consignor table
            consignor = Consignor.objects.filter(sender_mobile=uid).first()

            if consignor:
                # If it's in the Consignor table, fetch consignments based on sender_mobile
                user_branch = consignor.sender_mobile
                consignments = AddConsignment.objects.filter(sender_mobile=user_branch, consignment_status='Pending')

                # Loop over consignments and get their track_id
                for consignment in consignments:
                    track_id = consignment.track_id

                    # Fetch the trip details related to the consignment's track_id
                    trips = TripSheetPrem.objects.filter(LRno=track_id)

                    # If it's not already in grouped_userdata, initialize it
                    if track_id not in grouped_userdata:
                        grouped_userdata[track_id] = {
                            'route_from': consignment.route_from,
                            'route_to': consignment.route_to,
                            'sender_name': consignment.sender_name,
                            'sender_mobile': consignment.sender_mobile,
                            'receiver_name': consignment.receiver_name,
                            'receiver_mobile': consignment.receiver_mobile,
                            'total_cost': consignment.total_cost,
                            'pieces': 0,  # Aggregate pieces
                            'weight': consignment.weight,
                            'pay_status': consignment.pay_status,
                            'status': consignment.status,
                            'date': consignment.date,
                            'time': consignment.time,
                            'products': [],  # Aggregate product details
                            'trip_details': []  # Initialize a place to store trip details
                        }

                    # Aggregate total pieces at the consignment level
                    grouped_userdata[track_id]['pieces'] += consignment.pieces
                    # Concatenate product details without ID
                    product_detail = consignment.desc_product
                    grouped_userdata[track_id]['products'].append(product_detail)

                    # Add trip details for this consignment
                    for trip in trips:
                        trip_detail = {
                            'Date': trip.Date,
                            'Time': trip.Time,
                            'VehicalNo': trip.VehicalNo,
                            'DriverName': trip.DriverName,  # Assuming this field exists
                            'DriverNumber': trip.DriverNumber,  # Assuming this field exists
                            'sender_name': consignment.sender_name,
                            'receiver_name': consignment.receiver_name,
                            'route_from': consignment.route_from,
                            'route_to': consignment.route_to,
                            'pieces': consignment.pieces,  # Correctly assign pieces here from the consignment
                            'products': consignment.desc_product,  # Assign product details here
                        }
                        grouped_userdata[track_id]['trip_details'].append(trip_detail)

            else:
                # If not found in Consignor, check the Consignee table
                consignee = Consignee.objects.filter(receiver_mobile=uid).first()

                if consignee:
                    # If it's in the Consignee table, fetch consignments based on receiver_mobile
                    user_branch = consignee.receiver_mobile
                    consignments = AddConsignment.objects.filter(receiver_mobile=user_branch,
                                                                 consignment_status='Pending')

                    # Loop over consignments and get their track_id
                    for consignment in consignments:
                        track_id = consignment.track_id

                        # Fetch the trip details related to the consignment's track_id
                        trips = TripSheetPrem.objects.filter(LRno=track_id)

                        # If it's not already in grouped_userdata, initialize it
                        if track_id not in grouped_userdata:
                            grouped_userdata[track_id] = {
                                'route_from': consignment.route_from,
                                'route_to': consignment.route_to,
                                'sender_name': consignment.sender_name,
                                'sender_mobile': consignment.sender_mobile,
                                'receiver_name': consignment.receiver_name,
                                'receiver_mobile': consignment.receiver_mobile,
                                'total_cost': consignment.total_cost,
                                'pieces': 0,  # Aggregate pieces
                                'weight': consignment.weight,
                                'pay_status': consignment.pay_status,
                                'status': consignment.status,
                                'date': consignment.date,
                                'time': consignment.time,
                                'products': [],  # Aggregate product details
                                'trip_details': []  # Initialize a place to store trip details
                            }

                        # Aggregate total pieces at the consignment level
                        grouped_userdata[track_id]['pieces'] += consignment.pieces
                        # Concatenate product details without ID
                        product_detail = consignment.desc_product
                        grouped_userdata[track_id]['products'].append(product_detail)

                        # Add trip details for this consignment
                        for trip in trips:
                            trip_detail = {
                                'Date': trip.Date,
                                'Time': trip.Time,
                                'VehicalNo': trip.VehicalNo,
                                'DriverName': trip.DriverName,  # Assuming this field exists
                                'DriverNumber': trip.DriverNumber,  # Assuming this field exists
                                'sender_name': consignment.sender_name,
                                'receiver_name': consignment.receiver_name,
                                'route_from': consignment.route_from,
                                'route_to': consignment.route_to,
                                'pieces': consignment.pieces,  # Correctly assign pieces here from the consignment
                                'products': consignment.desc_product,  # Assign product details here
                            }
                            grouped_userdata[track_id]['trip_details'].append(trip_detail)

                else:
                    # If email is not found in either table, return empty
                    consignments = []
                    trips = []

        except ObjectDoesNotExist:
            # Handle case if neither Consignor nor Consignee exists for the uid
            pass

    # Convert the list of product details to a single string
    for track_id, details in grouped_userdata.items():
        details['products'] = ', '.join(details['products'])

    # Render the template with consignment and trip data
    return render(request, 'customerConsignment.html', {'grouped_userdata': grouped_userdata})


from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.core.exceptions import ObjectDoesNotExist
from .models import AddConsignment, Branch  # Adjust the import paths to your project structure


def downloadInvoice(request, track_id):
    grouped_userdata = {}
    copy_types = []
    totalqty = 0  # Initialize total quantity
    branchdetails = None  # Initialize branchdetails as None in case it's not found

    try:
        # Filter consignments by track_id
        consignments = AddConsignment.objects.filter(track_id=track_id)

        # Get the first consignment's branch (assuming all consignments for a track_id have the same branch)
        first_consignment = consignments.first()

        if first_consignment:
            bname = first_consignment.branch  # Access the branch from the first consignment

            if bname:
                try:
                    # Get branch details from the Branch model based on the branch name (as a string)
                    branchdetails = Branch.objects.get(companyname=bname)
                except Branch.DoesNotExist:
                    branchdetails = None  # If branch doesn't exist, set branchdetails to None

        # Loop over each consignment item to gather details individually
        for consignment in consignments:
            if consignment.track_id not in grouped_userdata:
                # Initialize data structure for each track_id
                grouped_userdata[consignment.track_id] = {
                    field.name: getattr(consignment, field.name) for field in AddConsignment._meta.fields
                }
                grouped_userdata[consignment.track_id]['consignment_list'] = []  # To store individual products

            # Add each consignment's product details as a separate entry
            consignment_details = {
                'pieces': consignment.pieces,
                'sender_name': consignment.sender_name,
                'sender_address': consignment.sender_address,
                'sender_GST': consignment.sender_GST,
                'sender_mobile': consignment.sender_mobile,
                'receiver_name': consignment.receiver_name,
                'receiver_address': consignment.receiver_address,
                'receiver_GST': consignment.receiver_GST,
                'receiver_mobile': consignment.receiver_mobile,
                'desc_product': consignment.desc_product,
                'weight': consignment.weight,
                'weightAmt': consignment.weightAmt,
                'prod_invoice': consignment.prod_invoice,
                'prod_price': consignment.prod_price,
                'total_cost': consignment.total_cost,
                'freight': consignment.freight,
                'hamali': consignment.hamali,
                'door_charge': consignment.door_charge,
                'st_charge': consignment.st_charge,
                'balance': consignment.balance,
                'consignment_status': consignment.consignment_status,
                'eway_bill': consignment.eway_bill,
                'delivery': consignment.delivery,
                'date': consignment.date,
                'time': consignment.time,
            }

            # Append consignment details for this track_id
            grouped_userdata[consignment.track_id]['consignment_list'].append(consignment_details)
            totalqty += consignment.pieces  # Sum up the pieces for total quantity

            if consignment.copy_type not in copy_types:
                copy_types.append(consignment.copy_type)

    except ObjectDoesNotExist:
        grouped_userdata = {}

    # Render the HTML content from the template
    html_content = render_to_string('downloadInvoice.html', {
        'grouped_userdata': grouped_userdata,
        'branchdetails': branchdetails,
        'copy_types': ', '.join(copy_types),
        'totalqty': totalqty
    })

    # Create the HTTP response for the PDF download
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{track_id}.pdf"'

    # Convert HTML to PDF using xhtml2pdf
    pisa_status = pisa.CreatePDF(html_content, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response


from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone


# Endpoint to check if location sharing is active
def check_location_sharing_status(request):
    phone_number = request.GET.get("phone_number", "")
    if not phone_number:
        return JsonResponse({"status": "error", "message": "Phone number is required."})

    try:
        driver = Driver.objects.get(phone_number=phone_number)
        location = DriverLocation.objects.filter(driver=driver, location_sharing_active=True).last()

        if location:
            return JsonResponse({"status": "success", "location_sharing_active": True})
        else:
            return JsonResponse({"status": "success", "location_sharing_active": False})

    except Driver.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Driver not found."})


# Endpoint to start location sharing
@csrf_exempt
def start_location_sharing(request, phone_number):
    try:
        driver = Driver.objects.get(phone_number=phone_number)
    except Driver.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Driver not found."})

    # Activate location sharing
    DriverLocation.objects.create(
        driver=driver,
        latitude=0,  # Placeholder for initial location
        longitude=0,  # Placeholder for initial location
        timestamp=timezone.now(),
        location_sharing_active=True
    )

    return JsonResponse({"status": "success", "message": "Location sharing started."})


# Endpoint to stop location sharing
@csrf_exempt
def stop_location_sharing(request, phone_number):
    try:
        driver = Driver.objects.get(phone_number=phone_number)
    except Driver.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Driver not found."})

    # Deactivate location sharing by updating the latest location entry
    last_location = DriverLocation.objects.filter(driver=driver, location_sharing_active=True).last()
    if last_location:
        last_location.location_sharing_active = False
        last_location.save()

    return JsonResponse({"status": "success", "message": "Location sharing stopped."})


# Endpoint to update the driver's location
@csrf_exempt
def update_location(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            phone_number = data.get("phone_number")
            latitude = data.get("latitude")
            longitude = data.get("longitude")
            timestamp = data.get("timestamp")

            # Check if required data is missing
            if not phone_number or not latitude or not longitude or not timestamp:
                return JsonResponse({"status": "error", "message": "Missing required fields."})

            # Validate the latitude and longitude values
            if not isinstance(latitude, (int, float)) or not isinstance(longitude, (int, float)):
                return JsonResponse({"status": "error", "message": "Invalid latitude or longitude format."})

            if latitude == 0 or longitude == 0:
                return JsonResponse(
                    {"status": "error", "message": "Invalid location (latitude or longitude cannot be zero)."})

            # Check if the driver exists
            try:
                driver = Driver.objects.get(phone_number=phone_number)
            except Driver.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Driver not found."})

            # Save or update the driver's location in the DriverLocation table
            location, created = DriverLocation.objects.update_or_create(
                driver=driver,
                defaults={
                    "latitude": latitude,
                    "longitude": longitude,
                    "timestamp": timestamp,
                    "location_sharing_active": True,  # Ensure sharing is active
                }
            )

            return JsonResponse({"status": "success", "message": "Location updated successfully."})

        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "Invalid JSON format."})

    return JsonResponse({"status": "error", "message": "Invalid request method."})


def location(request):
    uid = request.session.get('username')
    return render(request, 'update_location.html', {'uid': uid})


def track_driver(request):
    driver = None
    location = None

    phone_number = request.GET.get('phone_number')  # Get the phone number from the query parameters

    if phone_number:
        # Fetch the driver using the phone number
        driver = get_object_or_404(Driver, phone_number=phone_number)

        # Try to get the latest location or return None if no location exists
        location = DriverLocation.objects.filter(driver=driver).order_by('-timestamp').first()

    # Pass the driver and location to the template
    return render(request, 'track_driver.html', {
        'driver': driver,
        'location': location,
    })


def admintrack_driver(request):
    driver = None
    location = None

    phone_number = request.GET.get('phone_number')  # Get the phone number from the query parameters

    if phone_number:
        # Fetch the driver using the phone number
        driver = get_object_or_404(Driver, phone_number=phone_number)

        # Try to get the latest location or return None if no location exists
        location = DriverLocation.objects.filter(driver=driver).order_by('-timestamp').first()

    # Pass the driver and location to the template
    return render(request, 'admintrack_driver.html', {
        'driver': driver,
        'location': location,
    })


def home(request):
    return render(request, 'home.html')


def transporter(request):
    if request.method == "POST":
        companyname = request.POST.get('companyname')
        phonenumber = request.POST.get('phonenumber')

        gst = request.POST.get('gst')
        address = request.POST.get('address')

        if Transporter.objects.filter(companyname=companyname).exists():
            messages.error(request, 'companyname (email) already exists.')
            return render(request, 'transporter.html')

        # If the email does not exist, create the branch and login records
        Transporter.objects.create(
            companyname=companyname,
            phonenumber=phonenumber,
            gst=gst,
            address=address,
        )

        messages.success(request, 'Branch created successfully.')
    return render(request, 'transporter.html')


def view_transporter(request):
    data = Transporter.objects.all()
    return render(request, 'view_transporter.html', {'data': data})


def edit_transporter(request, pk):
    data = Transporter.objects.filter(id=pk).first()  # Retrieve a single object or None

    if request.method == "POST":
        companyname = request.POST.get('companyname')
        phonenumber = request.POST.get('phonenumber')
        gst = request.POST.get('gst')
        address = request.POST.get('address')

        # Update the object
        data.companyname = companyname
        data.phonenumber = phonenumber
        data.gst = gst
        data.address = address
        data.save()
        # Update the Login record using the original staffPhone
        # Redirect to a different URL after successful update
        base_url = reverse('view_transporter')
        return redirect(base_url)

    return render(request, 'edit_transporter.html', {'data': data})


def transporter_delete(request, pk):
    udata = Transporter.objects.get(id=pk)
    udata.delete()
    base_url = reverse('view_transporter')
    return redirect(base_url)


def fullLoad(request):
    if request.method == "POST":
        route_from = request.POST.get('route_from')
        route_to = request.POST.get('route_to')
        date = request.POST.get('date')
        rate = request.POST.get('rate')
        tone = request.POST.get('tone')
        total = request.POST.get('total')
        vehicleNo = request.POST.get('vehicleNo')
        driverName = request.POST.get('driver')
        transport = request.POST.get('transport')
        advance = request.POST.get('advance')
        goods = request.POST.get('goods')
        remark = request.POST.get('remark')
        balance = request.POST.get('balance')
        haltAmt = request.POST.get('haltAmt')
        dedicationReason = request.POST.get('dedicationReason')
        dedicationAmt = request.POST.get('dedicationAmt')
        finalBalance = request.POST.get('finalBalance')

        # Determine the payment status based on the balance
        status = 'Paid' if float(balance) == 0 else 'Pending'

        # Get the last track_id and increment it
        last_loadId = Fullload.objects.aggregate(Max('loadId'))['loadId__max']
        loadId = int(last_loadId) + 1 if last_loadId else 1000
        con_id = str(loadId)

        # Create the Fullload object
        Fullload.objects.create(
            route_from=route_from,
            route_to=route_to,
            date=date,
            rate=rate,
            tone=tone,
            total=total,
            vehicleNo=vehicleNo,
            driverName=driverName,
            transport=transport,
            advance=advance,
            paidAmt=advance,
            goods=goods,
            remark=remark,
            balance=balance,
            loadId=con_id,
            haltAmt=haltAmt,
            dedicationReason=dedicationReason,
            finalBalance=finalBalance,
            dedicationAmt=dedicationAmt,
            status=status  # Save the status (Paid or Pending)
        )

        messages.success(request, 'Full Load created successfully.')

    return render(request, 'fullLoad.html')


def fullLoadList(request):
    grouped_userdata = {}

    try:
        from_date_str = request.POST.get('fromdate')
        to_date_str = request.POST.get('todate')

        transport = request.POST.get('transport')
        driver = request.POST.get('driver')
        vehicleno = request.POST.get('vehicleno')


        # Parse dates
        from_date = parse_date(from_date_str) if from_date_str else None
        to_date = parse_date(to_date_str) if to_date_str else None


        # Start building the query
        consignments = Fullload.objects.all()

        if transport:
            consignments = consignments.filter(transport=transport)
        if vehicleno:
            consignments = consignments.filter(vehicleNo=vehicleno)
        if driver:
            consignments = consignments.filter(driverName=driver)

        if from_date and to_date:
            consignments = consignments.filter(date__range=(from_date, to_date))
        elif from_date:
            consignments = consignments.filter(date__gte=from_date)
        elif to_date:
            consignments = consignments.filter(date__lte=to_date)

        # Group consignments by track_id and concatenate product details
        for consignment in consignments:
            loadId = consignment.loadId
            if loadId not in grouped_userdata:
                grouped_userdata[loadId] = {
                    'id': consignment.id,
                    'route_from': consignment.route_from,
                    'route_to': consignment.route_to,
                    'date': consignment.date,
                    'transport': consignment.transport,
                    'vehicleNo': consignment.vehicleNo,
                    'rate': consignment.rate,
                    'tone': consignment.tone,
                    'total': consignment.total,
                    'goods': consignment.goods,
                    'advance': consignment.advance,
                    'balance': consignment.balance,
                    'finalBalance': consignment.finalBalance,
                    'remark': consignment.remark,
                    'status': consignment.status,
                    'paidAmt': consignment.paidAmt,
                    'driverName': consignment.driverName,
                    'haltAmt': consignment.haltAmt,
                    'dedicationAmt': consignment.dedicationAmt,
                    'dedicationReason': consignment.dedicationReason,
                }

    except ObjectDoesNotExist:
        # In case of staff or branch not found, return an empty set
        grouped_userdata = {}
    return render(request, 'fullLoadList.html',{'grouped_userdata':grouped_userdata})

def fullLoadBlancePage(request):
    grouped_userdata = {}

    try:
        from_date_str = request.POST.get('fromdate')
        to_date_str = request.POST.get('todate')

        transport = request.POST.get('transport')
        driver = request.POST.get('driver')
        vehicleno = request.POST.get('vehicleno')


        # Parse dates
        from_date = parse_date(from_date_str) if from_date_str else None
        to_date = parse_date(to_date_str) if to_date_str else None


        # Start building the query
        consignments = Fullload.objects.filter(status='Pending')

        if transport:
            consignments = consignments.filter(transport=transport)
        if vehicleno:
            consignments = consignments.filter(vehicleNo=vehicleno)
        if driver:
            consignments = consignments.filter(driverName=driver)
        if from_date and to_date:
            consignments = consignments.filter(date__range=(from_date, to_date))
        elif from_date:
            consignments = consignments.filter(date__gte=from_date)
        elif to_date:
            consignments = consignments.filter(date__lte=to_date)

        # Group consignments by track_id and concatenate product details
        for consignment in consignments:
            loadId = consignment.loadId
            if loadId not in grouped_userdata:
                grouped_userdata[loadId] = {
                    'route_from': consignment.route_from,
                    'route_to': consignment.route_to,
                    'date': consignment.date,
                    'transport': consignment.transport,
                    'vehicleNo': consignment.vehicleNo,
                    'rate': consignment.rate,
                    'tone': consignment.tone,
                    'total': consignment.total,
                    'goods': consignment.goods,
                    'advance': consignment.advance,
                    'paidAmt': consignment.paidAmt,
                    'dedicationAmt': consignment.dedicationAmt,
                    'dedicationReason': consignment.dedicationReason,
                    'paidAmt': consignment.paidAmt,
                    'finalBalance': consignment.finalBalance,
                    'balance': consignment.balance,
                    'remark': consignment.remark,
                    'status': consignment.status,
                }

    except ObjectDoesNotExist:
        # In case of staff or branch not found, return an empty set
        grouped_userdata = {}
    return render(request,'fullLoadBlancePage.html',{'grouped_userdata':grouped_userdata})

from decimal import Decimal
from django.shortcuts import get_object_or_404
from django.utils.timezone import now
from django.http import JsonResponse


def update_payment_status(request):
    if request.method == "POST":
        load_id = request.POST.get('loadId')  # From the hidden input in the modal
        amount = request.POST.get('amount')  # Amount input in the modal
        dedicationAmt = request.POST.get('dedicationamt')  # Amount input in the modal
        dedicationamtReason = request.POST.get('dedicationamtReason')  # Reason for the dedication amount
        remark = request.POST.get('remark')  # Remark input in the modal

        try:
            # Validate amount
            if not amount:
                return JsonResponse({'success': False, 'message': 'Amount is required.'})

            # Convert amount to float
            try:
                amount = float(amount)  # Convert to float
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Invalid amount format.'})

            # Convert dedicationAmt to float (handle empty values as 0)
            dedicationAmt = float(dedicationAmt) if dedicationAmt else 0

            # Get the Fullload entry
            full_load = get_object_or_404(Fullload, loadId=load_id)

            # Ensure fields exist and perform the update
            if not hasattr(full_load, 'paidAmt') or not hasattr(full_load, 'balance'):
                return JsonResponse({'success': False, 'message': 'Model does not have the required fields.'})

            # Update PaidAmt and subtract the amount from Balance
            full_load.paidAmt = (full_load.paidAmt or 0) + amount
            full_load.dedicationAmt = (full_load.dedicationAmt or 0) + dedicationAmt
            full_load.balance = (full_load.balance or 0) - amount - dedicationAmt
            full_load.finalBalance = (full_load.finalBalance or 0) - amount - dedicationAmt
            full_load.remark = remark

            # Update dedicationamtReason in Fullload
            full_load.dedicationamtReason = dedicationamtReason  # Update the field with the provided reason

            # Check if balance is zero, update status to "Paid"
            if full_load.finalBalance == 0:
                full_load.status = "Paid"

            full_load.save()

            # Create a new Balance entry
            Balance.objects.create(
                full_load=full_load,
                loadId=load_id,
                change_amount=amount,
                dedicationAmt=dedicationAmt,
                remark=remark,
                dedicationamtReason=dedicationamtReason,
                date=now().date()
            )

            return JsonResponse({'success': True, 'message': 'Payment status updated successfully.'})

        except Exception as e:
            # Return a detailed error message for debugging
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


def fullLoadExpenses(request):
    if request.method == "POST":
        date = request.POST.get('date')
        amount = request.POST.get('amount')
        name = request.POST.get('name')
        reason = request.POST.get('reason')
        desc = request.POST.get('desc')

        Fullloadexpenses.objects.create(
            Date=date,
            Amount=amount,
            staffname=name,
            Reason=reason,
            desc=desc,
        )
    return render(request,'fullLoadExpenses.html')

def fullLoadExpensesList(request):
    # Initialize the queryset
    data = Fullloadexpenses.objects.all()

    # Get POST parameters
    from_date_str = request.POST.get('fromdate')
    to_date_str = request.POST.get('todate')

    # Parse dates
    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None

    # Apply filters based on dates
    if from_date and to_date:
        data = data.filter(Date__range=(from_date, to_date))
    elif from_date:
        data = data.filter(Date__gte=from_date)
    elif to_date:
        data = data.filter(Date__lte=to_date)

    return render(request, 'fullLoadExpensesList.html', {'data': data})

from django.db.models import Sum


def fullLoadProfit(request):
    # Get the from_date and to_date from the request (if provided)
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None

    # Query all consignments and expenses
    consignments = Fullload.objects.all()
    expenses = Fullloadexpenses.objects.all()

    # Filter by date range if provided
    if from_date and to_date:
        consignments = consignments.filter(date__range=[from_date, to_date])
        expenses = expenses.filter(Date__range=[from_date, to_date])

    # Track already processed loadIds
    processed_loadIds = set()

    # This list will store the unique consignments
    unique_consignments = []

    # Iterate over all consignments to ensure only unique loadId costs are added
    for consignment in consignments:
        loadId = consignment.loadId
        # If loadId is not already processed, add it to the unique list
        if loadId not in processed_loadIds:
            processed_loadIds.add(loadId)
            unique_consignments.append(consignment)

    # Now group by date and branch, summing the advance and paidAmt to get the total
    consignments_grouped = (
        Fullload.objects.filter(id__in=[c.id for c in unique_consignments])
        .values('date')
        .annotate(
            total=Sum('advance') + Sum('paidAmt')  # Sum advance and paidAmt as total
        )
        .order_by('date')
    )

    # Group expenses by date and calculate the total Amount for each group
    expenses_grouped = expenses.values('Date').annotate(
        total_amount=Sum('Amount')
    ).order_by('Date')

    # Calculate grand totals for consignments using sum of advance and paidAmt
    grand_total_consignment = sum(
        item['total'] for item in consignments_grouped
    )

    # Calculate grand total of expenses
    grand_total_expenses = sum(
        item['total_amount'] for item in expenses_grouped
    )

    # Calculate combined grand total (consignments + expenses)
    combined_grand_total = grand_total_consignment + grand_total_expenses

    # Calculate profit or loss
    total_balance = grand_total_consignment - grand_total_expenses

    # Set profit and loss
    profit = total_balance if total_balance > 0 else 0
    loss = abs(total_balance) if total_balance < 0 else 0

    # Pass the grouped data and totals to the template
    return render(request, 'fullLoadProfit.html', {
        'consignments': consignments_grouped,
        'expenses': expenses_grouped,
        'grand_total_consignment': grand_total_consignment,
        'grand_total_expenses': grand_total_expenses,
        'combined_grand_total': combined_grand_total,
        'profit': profit,
        'loss': loss,
        'from_date': from_date_str,
        'to_date': to_date_str,
    })

def fullLoadInvoice(request,pk):
    data = Fullload.objects.filter(loadId=pk)
    return render(request,'fullLoadInvoice.html',{'data':data})



def fullLoadEdit(request, id):
    data = Fullload.objects.filter(id=id).first()  # Retrieve a single object or None

    if request.method == "POST":
        route_from = request.POST.get('route_from')
        route_to = request.POST.get('route_to')
        date = request.POST.get('date')
        rate = request.POST.get('rate')
        tone = request.POST.get('tone')
        total = request.POST.get('total')
        vehicleNo = request.POST.get('vehicleNo')
        driverName = request.POST.get('driver')
        transport = request.POST.get('transport')
        advance = request.POST.get('advance')
        goods = request.POST.get('goods')
        remark = request.POST.get('remark')
        balance = request.POST.get('balance')
        haltAmt = request.POST.get('haltAmt')
        dedicationReason = request.POST.get('dedicationReason')
        dedicationAmt = request.POST.get('dedicationAmt')
        finalBalance = request.POST.get('finalBalance')

        # Update the object
        data.route_from = route_from
        data.route_to = route_to
        data.date = date
        data.rate = rate
        data.tone = tone
        data.total = total
        data.vehicleNo = vehicleNo
        data.driverName = driverName
        data.transport = transport
        data.advance = advance
        data.paidAmt = advance
        data.goods = goods
        data.remark = remark
        data.balance = balance
        data.haltAmt = haltAmt
        data.dedicationReason = dedicationReason
        data.dedicationAmt = dedicationAmt
        data.finalBalance = finalBalance
        data.save()

        # Redirect to a different URL after successful update
        base_url = reverse('fullLoadList')
        return redirect(base_url)

    return render(request, 'fullLoadEdit.html', {'data': data})
